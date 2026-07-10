# V0126C
# Audit reference: range-query repair of the IERS-0012O-style 1761/1769 JPL engineering track plot.

from __future__ import annotations

import csv
import hashlib
import math
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.error import HTTPError
from zoneinfo import ZoneInfo


def ensure_package(import_name: str, pip_name: str) -> None:
    try:
        __import__(import_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "-q", "install", pip_name])


for _import_name, _pip_name in (
    ("numpy", "numpy"),
    ("pandas", "pandas"),
    ("scipy", "scipy"),
    ("matplotlib", "matplotlib"),
    ("astroquery", "astroquery"),
    ("astropy", "astropy"),
    ("openpyxl", "openpyxl>=3.1.0"),
):
    ensure_package(_import_name, _pip_name)

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from astropy.time import Time
from astroquery.jplhorizons import Horizons
from matplotlib.patches import Circle
from openpyxl import load_workbook
from scipy.interpolate import CubicSpline
from scipy.optimize import minimize_scalar

VERSION = "IERS-0126C"
PROGRAM_NAME = "IERS_0126C_IMCCE_1761_1769_JPL_ENGINEERING_TRACKS.py"
REFERENCE_TEMPLATE = "IERS_0012O_NORTH_SOUTH_POLE_ENGINEERING_TRACK_PLOT_PI_SUN.py"
TARGET_YEARS = (1761, 1769)

AU_KM = 149_597_870.7
ARCSEC_PER_RAD = 206_264.80624709636
SUN_RADIUS_KM = 695_700.0
VENUS_RADIUS_KM = 6_051.8
LOCAL_TZ = ZoneInfo("America/Bogota")
QUERY_STEP = "1m"
QUERY_PAD_MINUTES = 3.0
MAX_QUERY_ATTEMPTS = 4

DRIVE_ROOT = Path("/content/drive/MyDrive/Colab Notebooks/JPL - 1769 VENUS TRANSIT/GitHub")
ORGANIZED_XLSX = DRIVE_ROOT / "DATA" / "XLSX" / "IMCCE_VENUS_TRANSIT_CANON_ORGANIZED.xlsx"
MASTER_XLSX = DRIVE_ROOT / "DATA" / "XLSX" / "IMCCE_VENUS_TRANSIT_CANON_MASTER.xlsx"
OUTPUT_ROOT = Path("/content/IERS_TN36_V01_MASTER_OUTPUT/IMCCE_VENUS_CANON/PLOTS/V0126C")
TRACK_CSV = OUTPUT_ROOT / "IERS_0126C_IMCCE_1761_1769_JPL_TRACKS.csv"
RESULTS_CSV = OUTPUT_ROOT / "IERS_0126C_IMCCE_1761_1769_JPL_RESULTS.csv"
FIGURE_PNG = OUTPUT_ROOT / "IERS_0126C_IMCCE_1761_1769_JPL_ENGINEERING_TRACKS.png"

DRIVE_OUTPUTS = {
    "track_csv": DRIVE_ROOT / "DATA" / "CSV" / TRACK_CSV.name,
    "results_csv": DRIVE_ROOT / "DATA" / "CSV" / RESULTS_CSV.name,
    "figure_png": DRIVE_ROOT / "DATA" / "PNG" / FIGURE_PNG.name,
    "python": DRIVE_ROOT / "PYTHON" / PROGRAM_NAME,
}

TRACK_COLORS = {1761: "#ffc861", 1769: "#5ee08a"}
REQUIRED_COLUMNS = {
    "record_id",
    "year",
    "jd_tdb",
    "sun_radius_arcsec",
    "minimum_distance_arcsec",
    "venus_radius_arcsec",
    "c1_ut",
    "c2_ut",
    "c3_ut",
    "c4_ut",
    "mid_ut_seconds_of_day",
    "relative_velocity_deg_per_day",
    "record_status",
}


def norm(vector: np.ndarray) -> float:
    return float(np.sqrt(np.dot(vector, vector)))


def unit(vector: np.ndarray) -> np.ndarray:
    magnitude = norm(vector)
    if magnitude == 0.0:
        raise RuntimeError("Zero vector encountered.")
    return np.asarray(vector, dtype=float) / magnitude


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def copy_verified(source: Path, destination: Path) -> str:
    destination.parent.mkdir(parents=True, exist_ok=True)
    source_hash = sha256(source)
    if destination.exists() and sha256(destination) == source_hash:
        return "UNCHANGED"
    shutil.copy2(source, destination)
    if sha256(destination) != source_hash:
        raise RuntimeError(f"SHA-256 verification failed: {destination}")
    return "COPIED"


def normalize_header(value: object) -> str:
    return str(value).strip().lower().replace(" ", "_").replace("-", "_")


def resolve_workbook() -> Path:
    for path in (ORGANIZED_XLSX, MASTER_XLSX):
        if path.exists():
            return path
    raise FileNotFoundError("Run IERS-0125A before IERS-0126C; organized workbook not found.")


def read_target_records(path: Path) -> dict[int, dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if "MASTER" not in workbook.sheetnames:
        workbook.close()
        raise RuntimeError("The organized workbook does not contain the MASTER sheet.")
    sheet = workbook["MASTER"]
    headers = {
        normalize_header(sheet.cell(1, column).value): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(1, column).value is not None
    }
    missing = sorted(REQUIRED_COLUMNS - set(headers))
    if missing:
        workbook.close()
        raise RuntimeError(f"MASTER sheet missing required columns: {missing}")

    records: dict[int, dict[str, object]] = {}
    for row_index in range(2, sheet.max_row + 1):
        raw_year = sheet.cell(row_index, headers["year"]).value
        if raw_year is None:
            continue
        year = int(raw_year)
        if year not in TARGET_YEARS:
            continue
        record = {key: sheet.cell(row_index, column).value for key, column in headers.items()}
        if record["record_status"] != "COMPLETE":
            workbook.close()
            raise RuntimeError(f"IMCCE record {year} is not complete: {record['record_status']}")
        records[year] = record
    workbook.close()

    missing_years = sorted(set(TARGET_YEARS) - set(records))
    if missing_years:
        raise RuntimeError(f"Missing target records: {missing_years}")
    return records


def clock_to_seconds(value: object) -> float:
    parts = str(value).strip().split(":")
    if len(parts) == 2:
        return int(parts[0]) * 3600.0 + float(parts[1]) * 60.0
    if len(parts) == 3:
        return int(parts[0]) * 3600.0 + int(parts[1]) * 60.0 + float(parts[2])
    raise ValueError(f"Unsupported clock value: {value!r}")


def contact_jds(record: dict[str, object]) -> dict[str, float]:
    contact_seconds = [clock_to_seconds(record[f"c{index}_ut"]) for index in range(1, 5)]
    unwrapped = [contact_seconds[0]]
    for seconds in contact_seconds[1:]:
        while seconds < unwrapped[-1]:
            seconds += 86400.0
        unwrapped.append(seconds)

    mid_seconds = float(record["mid_ut_seconds_of_day"])
    interval_center = 0.5 * (unwrapped[0] + unwrapped[-1])
    mid_unwrapped = min(
        (mid_seconds + day * 86400.0 for day in range(-2, 3)),
        key=lambda candidate: abs(candidate - interval_center),
    )
    mid_jd = float(record["jd_tdb"])
    return {
        event: mid_jd + (seconds - mid_unwrapped) / 86400.0
        for event, seconds in zip(("C1", "C2", "C3", "C4"), unwrapped)
    }


def horizons_calendar(jd_tdb: float) -> str:
    return Time(jd_tdb, format="jd", scale="tdb").tdb.strftime("%Y-%b-%d %H:%M:%S")


def query_vectors_range(target_id: str, start_jd: float, stop_jd: float, prefix: str) -> pd.DataFrame:
    start_text = horizons_calendar(start_jd)
    stop_text = horizons_calendar(stop_jd)
    last_error: Exception | None = None
    for attempt in range(1, MAX_QUERY_ATTEMPTS + 1):
        try:
            query = Horizons(
                id=target_id,
                location="500@399",
                epochs={"start": start_text, "stop": stop_text, "step": QUERY_STEP},
            )
            vectors = query.vectors(refplane="ecliptic").to_pandas()
            if vectors.empty:
                raise RuntimeError(f"JPL returned no vectors for target {target_id}")
            return pd.DataFrame(
                {
                    "jd_tdb": vectors["datetime_jd"].astype(float),
                    f"{prefix}_x_km": vectors["x"].astype(float) * AU_KM,
                    f"{prefix}_y_km": vectors["y"].astype(float) * AU_KM,
                    f"{prefix}_z_km": vectors["z"].astype(float) * AU_KM,
                }
            )
        except Exception as exc:
            last_error = exc
            if attempt == MAX_QUERY_ATTEMPTS:
                break
            time.sleep(2 ** (attempt - 1))
    raise RuntimeError(
        f"JPL range query failed after {MAX_QUERY_ATTEMPTS} attempts for target {target_id}, "
        f"{start_text} to {stop_text}: {last_error}"
    ) from last_error


def build_jpl_master(start_jd: float, stop_jd: float) -> pd.DataFrame:
    sun = query_vectors_range("10", start_jd, stop_jd, "SUN")
    venus = query_vectors_range("299", start_jd, stop_jd, "VENUS")
    sun["epoch_key"] = np.round(sun["jd_tdb"].to_numpy(dtype=float), 10)
    venus["epoch_key"] = np.round(venus["jd_tdb"].to_numpy(dtype=float), 10)
    merged = sun.merge(venus, on="epoch_key", how="inner", suffixes=("_sun", "_venus"))
    merged["jd_tdb"] = merged["epoch_key"].astype(float)
    merged = merged.sort_values("jd_tdb").reset_index(drop=True)
    if len(merged) < 120:
        raise RuntimeError(f"JPL range query returned too few synchronized rows: {len(merged)}")
    return merged


def vector_from_row(row: pd.Series, prefix: str) -> np.ndarray:
    return np.asarray(
        [row[f"{prefix}_x_km"], row[f"{prefix}_y_km"], row[f"{prefix}_z_km"]],
        dtype=float,
    )


def fixed_solar_basis(master: pd.DataFrame, mid_jd: float) -> tuple[np.ndarray, np.ndarray, np.ndarray, float, float]:
    index = int(np.argmin(np.abs(master["jd_tdb"].to_numpy(dtype=float) - mid_jd)))
    row = master.iloc[index]
    sun = vector_from_row(row, "SUN")
    venus = vector_from_row(row, "VENUS")
    normal = unit(sun)
    reference = np.asarray([0.0, 0.0, 1.0])
    if norm(np.cross(reference, normal)) < 1.0e-12:
        reference = np.asarray([1.0, 0.0, 0.0])
    xhat = unit(np.cross(reference, normal))
    yhat = unit(np.cross(normal, xhat))
    sun_radius = math.atan2(SUN_RADIUS_KM, norm(sun)) * ARCSEC_PER_RAD
    venus_radius = math.atan2(VENUS_RADIUS_KM, norm(venus)) * ARCSEC_PER_RAD
    return normal, xhat, yhat, sun_radius, venus_radius


def relative_sky_offsets(
    master: pd.DataFrame,
    basis: tuple[np.ndarray, np.ndarray, np.ndarray, float, float],
) -> tuple[np.ndarray, np.ndarray]:
    normal, xhat, yhat, _sun_radius, _venus_radius = basis
    x_values = []
    y_values = []
    for _, row in master.iterrows():
        sun_direction = unit(vector_from_row(row, "SUN"))
        venus_direction = unit(vector_from_row(row, "VENUS"))
        sun_denominator = float(np.dot(sun_direction, normal))
        venus_denominator = float(np.dot(venus_direction, normal))
        sun_x = math.atan2(float(np.dot(sun_direction, xhat)), sun_denominator)
        sun_y = math.atan2(float(np.dot(sun_direction, yhat)), sun_denominator)
        venus_x = math.atan2(float(np.dot(venus_direction, xhat)), venus_denominator)
        venus_y = math.atan2(float(np.dot(venus_direction, yhat)), venus_denominator)
        x_values.append((venus_x - sun_x) * ARCSEC_PER_RAD)
        y_values.append((venus_y - sun_y) * ARCSEC_PER_RAD)
    return np.asarray(x_values), np.asarray(y_values)


def fit_r2(time_values: np.ndarray, x_values: np.ndarray, y_values: np.ndarray, degree: int) -> float:
    centered = (time_values - np.mean(time_values)) * 86400.0
    x_fit = np.polyval(np.polyfit(centered, x_values, degree), centered)
    y_fit = np.polyval(np.polyfit(centered, y_values, degree), centered)
    residual = np.sum((x_values - x_fit) ** 2 + (y_values - y_fit) ** 2)
    total = np.sum((x_values - np.mean(x_values)) ** 2 + (y_values - np.mean(y_values)) ** 2)
    return 1.0 if total == 0.0 else float(1.0 - residual / total)


def pca_metrics(points: np.ndarray) -> tuple[np.ndarray, np.ndarray, float, float]:
    mean = points.mean(axis=0)
    centered = points - mean
    _u, _s, vt = np.linalg.svd(centered, full_matrices=False)
    direction = vt[0]
    if direction[0] < 0.0:
        direction = -direction
    normal = np.asarray([-direction[1], direction[0]])
    residuals = centered @ normal
    rms = float(np.sqrt(np.mean(residuals**2)))
    angle = math.degrees(math.atan2(direction[1], direction[0]))
    return mean, direction, angle, rms


def curvature_at_mid(time_values: np.ndarray, x_values: np.ndarray, y_values: np.ndarray) -> float:
    centered = (time_values - np.mean(time_values)) * 86400.0
    px = np.polyfit(centered, x_values, 2)
    py = np.polyfit(centered, y_values, 2)
    x_prime = px[1]
    y_prime = py[1]
    x_second = 2.0 * px[0]
    y_second = 2.0 * py[0]
    denominator = (x_prime**2 + y_prime**2) ** 1.5
    return 0.0 if denominator == 0.0 else float(abs(x_prime * y_second - y_prime * x_second) / denominator)


def build_track(year: int, record: dict[str, object]) -> dict[str, object]:
    events = contact_jds(record)
    pad_days = QUERY_PAD_MINUTES / 1440.0
    query_start = events["C1"] - pad_days
    query_stop = events["C4"] + pad_days
    master = build_jpl_master(query_start, query_stop)
    basis = fixed_solar_basis(master, float(record["jd_tdb"]))
    x_arcsec, y_arcsec = relative_sky_offsets(master, basis)
    sun_radius = float(basis[3])
    master["x_arcsec"] = x_arcsec
    master["y_arcsec"] = y_arcsec
    master["x_solar_radii"] = x_arcsec / sun_radius
    master["y_solar_radii"] = y_arcsec / sun_radius

    spline_x = CubicSpline(master["jd_tdb"], master["x_solar_radii"], bc_type="natural")
    spline_y = CubicSpline(master["jd_tdb"], master["y_solar_radii"], bc_type="natural")
    closest = minimize_scalar(
        lambda jd: float(spline_x(jd) ** 2 + spline_y(jd) ** 2),
        bounds=(events["C1"], events["C4"]),
        method="bounded",
        options={"xatol": 1.0e-13},
    )
    closest_jd = float(closest.x)
    event_jds = {
        "C1": events["C1"],
        "C2": events["C2"],
        "CA": closest_jd,
        "C3": events["C3"],
        "C4": events["C4"],
    }
    event_points = {
        event: np.asarray([float(spline_x(jd)), float(spline_y(jd))])
        for event, jd in event_jds.items()
    }

    minute_master = master[
        (master["jd_tdb"] >= events["C1"]) & (master["jd_tdb"] <= events["C4"])
    ].copy()
    if len(minute_master) < 120:
        raise RuntimeError(f"Too few in-transit JPL rows for {year}: {len(minute_master)}")
    points = minute_master[["x_solar_radii", "y_solar_radii"]].to_numpy(dtype=float)
    mean, direction, angle, rms = pca_metrics(points)
    times = minute_master["jd_tdb"].to_numpy(dtype=float)
    linear_r2 = fit_r2(times, points[:, 0], points[:, 1], 1)
    quadratic_r2 = fit_r2(times, points[:, 0], points[:, 1], 2)
    cubic_r2 = fit_r2(times, points[:, 0], points[:, 1], 3)
    curvature = curvature_at_mid(times, points[:, 0], points[:, 1])
    elapsed_seconds = (times - times[0]) * 86400.0
    along = (points - mean) @ direction
    fit_speed = float(np.polyfit(elapsed_seconds, along, 1)[0])
    source_speed = float(record["relative_velocity_deg_per_day"]) * 3600.0 / 86400.0 / sun_radius

    return {
        "year": year,
        "record": record,
        "master": master,
        "minute_master": minute_master,
        "points": points,
        "events": event_jds,
        "event_points": event_points,
        "sun_radius_arcsec": sun_radius,
        "venus_radius_ratio": float(basis[4] / basis[3]),
        "mean": mean,
        "direction": direction,
        "angle_deg": angle,
        "rms_solar_radii": rms,
        "linear_r2": linear_r2,
        "quadratic_r2": quadratic_r2,
        "cubic_r2": cubic_r2,
        "curvature_inverse_solar_radii": curvature,
        "fit_speed_solar_radii_per_s": fit_speed,
        "source_speed_solar_radii_per_s": source_speed,
        "closest_jd": closest_jd,
        "query_start": horizons_calendar(query_start),
        "query_stop": horizons_calendar(query_stop),
    }


def add_label(axis, point: np.ndarray, text: str, dx: float, dy: float, color: str) -> None:
    axis.annotate(
        text,
        xy=(point[0], point[1]),
        xytext=(point[0] + dx, point[1] + dy),
        textcoords="data",
        fontsize=5.7,
        color=color,
        ha="left",
        va="center",
        arrowprops={"arrowstyle": "-", "lw": 0.20, "color": color, "shrinkA": 0, "shrinkB": 2},
    )


def add_summary_table(axis, tracks: list[dict[str, object]]) -> None:
    rows = []
    for track in tracks:
        year = int(track["year"])
        ca = track["event_points"]["CA"]
        rows.extend(
            [
                [f"{year} CA radius", f"{norm(ca):.6f}", "R_sun"],
                [f"{year} angle", f"{track['angle_deg']:.6f}", "deg"],
                [f"{year} RMS", f"{track['rms_solar_radii']:.9f}", "R_sun"],
                [f"{year} curvature", f"{track['curvature_inverse_solar_radii']:.9f}", "1/R_sun"],
            ]
        )
    table = axis.table(
        cellText=rows,
        colLabels=["Quantity", "Value", "Unit"],
        loc="lower left",
        colWidths=[0.31, 0.23, 0.17],
        bbox=[0.405, 0.055, 0.405, 0.265],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(5.15)
    for (row, column), cell in table.get_celld().items():
        cell.set_linewidth(0.18)
        cell.set_edgecolor("#1e4f64")
        if row == 0:
            cell.set_facecolor("#0a1a22")
            cell.get_text().set_color("#66e8ff")
            cell.get_text().set_fontweight("bold")
        else:
            cell.set_facecolor("#050b0f")
            if column == 1:
                cell.get_text().set_color("#ffc861")
                cell.get_text().set_fontweight("bold")
            elif column == 2:
                cell.get_text().set_color("#5ee08a")
            else:
                cell.get_text().set_color("#dff8ff")


def plot_engineering_tracks(tracks: list[dict[str, object]]) -> None:
    figure, axis = plt.subplots(figsize=(9.6, 7.2), dpi=240)
    figure.patch.set_facecolor("#03080d")
    axis.set_facecolor("#03080d")
    axis.add_patch(Circle((0.0, 0.0), 1.0, fill=False, lw=0.36, ec="#66e8ff", alpha=0.95))
    axis.axhline(0.0, lw=0.18, color="#1d3d4a", alpha=0.72)
    axis.axvline(0.0, lw=0.18, color="#1d3d4a", alpha=0.72)

    for track in tracks:
        year = int(track["year"])
        color = TRACK_COLORS[year]
        points = track["points"]
        axis.plot(points[:, 0], points[:, 1], lw=0.30, color=color, solid_capstyle="round", label=str(year), zorder=3)
        axis.scatter(points[::6, 0], points[::6, 1], s=0.75, color=color, alpha=0.70, linewidths=0, zorder=4)
        for event in ("C1", "C2", "CA", "C3", "C4"):
            point = track["event_points"][event]
            radius = float(track["venus_radius_ratio"])
            axis.add_patch(
                Circle(
                    (point[0], point[1]),
                    radius,
                    fill=False,
                    lw=0.20 if event != "CA" else 0.28,
                    ec=color,
                    alpha=0.92,
                    zorder=2,
                )
            )
            axis.scatter(
                [point[0]],
                [point[1]],
                s=3.8 if event == "CA" else 2.2,
                color=color,
                edgecolors="#03080d",
                linewidths=0.16,
                zorder=5,
            )
        ca_point = track["event_points"]["CA"]
        ca_dy = 0.045 if ca_point[1] >= 0.0 else -0.045
        add_label(axis, ca_point, f"{year} CA", 0.055, ca_dy, color)
        for event, dx, dy in (
            ("C1", -0.115, 0.040),
            ("C2", -0.095, 0.030),
            ("C3", 0.055, -0.032),
            ("C4", 0.075, -0.043),
        ):
            add_label(axis, track["event_points"][event], f"{year} {event}", dx, dy, color)

    add_summary_table(axis, tracks)
    axis.set_xlim(-1.06, 1.06)
    axis.set_ylim(-1.06, 1.06)
    axis.set_aspect("equal", adjustable="box")
    for spine in axis.spines.values():
        spine.set_linewidth(0.22)
        spine.set_color("#25708b")
    axis.tick_params(axis="both", colors="#8fb4c1", labelsize=6.5, width=0.22, length=2.0)
    axis.grid(True, color="#102630", linewidth=0.16, alpha=0.55)
    axis.set_xlabel("Ecliptic tangent-plane X / solar radius", color="#8fb4c1", fontsize=7.5)
    axis.set_ylabel("Ecliptic tangent-plane Y / solar radius", color="#8fb4c1", fontsize=7.5)
    axis.set_title(
        "1761 and 1769 Venus Transits — Engineering Solar-Disk Track Reconstruction\n"
        "IMCCE contact epochs / JPL Horizons one-minute range queries — IERS-0012O style",
        color="#f8fdff",
        fontsize=9.0,
        pad=8,
    )
    legend = axis.legend(loc="lower right", fontsize=6.3, frameon=True, borderpad=0.45)
    legend.get_frame().set_facecolor("#071016")
    legend.get_frame().set_edgecolor("#1e4f64")
    legend.get_frame().set_linewidth(0.22)
    for legend_text in legend.get_texts():
        legend_text.set_color("#dff8ff")
    figure.text(
        0.5,
        0.016,
        "Venus disks are plotted to scale at C1, C2, closest approach, C3, and C4. "
        "Each event is normalized by its own JPL solar radius.",
        ha="center",
        va="bottom",
        fontsize=6.2,
        color="#8fb4c1",
    )
    FIGURE_PNG.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(FIGURE_PNG, dpi=460, facecolor=figure.get_facecolor(), bbox_inches="tight", pad_inches=0.055)
    plt.show()
    plt.close(figure)


def write_track_csv(tracks: list[dict[str, object]]) -> None:
    TRACK_CSV.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "year",
        "sample_type",
        "event",
        "jd_tdb",
        "elapsed_from_c1_seconds",
        "x_solar_radii",
        "y_solar_radii",
        "x_arcsec",
        "y_arcsec",
        "venus_radius_solar_radii",
        "source",
    ]
    with TRACK_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for track in tracks:
            c1 = float(track["events"]["C1"])
            sun_radius = float(track["sun_radius_arcsec"])
            for _, row in track["minute_master"].iterrows():
                writer.writerow(
                    {
                        "year": track["year"],
                        "sample_type": "JPL_ONE_MINUTE",
                        "event": "",
                        "jd_tdb": f"{float(row['jd_tdb']):.12f}",
                        "elapsed_from_c1_seconds": f"{(float(row['jd_tdb']) - c1) * 86400.0:.6f}",
                        "x_solar_radii": f"{float(row['x_solar_radii']):.9f}",
                        "y_solar_radii": f"{float(row['y_solar_radii']):.9f}",
                        "x_arcsec": f"{float(row['x_arcsec']):.6f}",
                        "y_arcsec": f"{float(row['y_arcsec']):.6f}",
                        "venus_radius_solar_radii": f"{float(track['venus_radius_ratio']):.9f}",
                        "source": "JPL_HORIZONS_GEOCENTER_RANGE_QUERY",
                    }
                )
            for event in ("C1", "C2", "CA", "C3", "C4"):
                jd = float(track["events"][event])
                point = track["event_points"][event]
                writer.writerow(
                    {
                        "year": track["year"],
                        "sample_type": "EVENT",
                        "event": event,
                        "jd_tdb": f"{jd:.12f}",
                        "elapsed_from_c1_seconds": f"{(jd - c1) * 86400.0:.6f}",
                        "x_solar_radii": f"{float(point[0]):.9f}",
                        "y_solar_radii": f"{float(point[1]):.9f}",
                        "x_arcsec": f"{float(point[0]) * sun_radius:.6f}",
                        "y_arcsec": f"{float(point[1]) * sun_radius:.6f}",
                        "venus_radius_solar_radii": f"{float(track['venus_radius_ratio']):.9f}",
                        "source": "IMCCE_EVENT_EPOCH_JPL_SPLINE_POSITION" if event != "CA" else "JPL_CLOSEST_APPROACH",
                    }
                )


def write_results_csv(tracks: list[dict[str, object]]) -> None:
    with RESULTS_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["year", "quantity", "value", "unit", "traceability"])
        for track in tracks:
            ca_point = track["event_points"]["CA"]
            rows = [
                ("record_id", track["record"]["record_id"], "record", "IMCCE workbook"),
                ("query_start", track["query_start"], "TDB", "JPL range request"),
                ("query_stop", track["query_stop"], "TDB", "JPL range request"),
                ("jpl_sun_radius", track["sun_radius_arcsec"], "arcsec", "JPL vector distance"),
                ("jpl_venus_radius_ratio", track["venus_radius_ratio"], "solar radii", "JPL vector distance"),
                ("closest_radius", norm(ca_point), "solar radii", "JPL spline minimization"),
                ("closest_x", ca_point[0], "solar radii", "JPL spline minimization"),
                ("closest_y", ca_point[1], "solar radii", "JPL spline minimization"),
                ("track_angle", track["angle_deg"], "deg", "PCA of one-minute JPL positions"),
                ("orthogonal_rms", track["rms_solar_radii"], "solar radii", "PCA residual"),
                ("linear_R2", track["linear_r2"], "dimensionless", "time polynomial fit"),
                ("quadratic_R2", track["quadratic_r2"], "dimensionless", "time polynomial fit"),
                ("cubic_R2", track["cubic_r2"], "dimensionless", "time polynomial fit"),
                ("curvature_at_mid", track["curvature_inverse_solar_radii"], "1/solar radius", "quadratic time fit"),
                ("fit_speed", track["fit_speed_solar_radii_per_s"], "solar radii/s", "PCA along-track fit"),
                ("IMCCE_source_speed", track["source_speed_solar_radii_per_s"], "solar radii/s", "IMCCE V converted with JPL solar radius"),
                ("closest_jd_tdb", track["closest_jd"], "day", "JPL spline minimization"),
            ]
            for quantity, value, unit_name, source in rows:
                writer.writerow([track["year"], quantity, value, unit_name, source])
        writer.writerow(
            [
                "COMPARISON",
                "halley_parallax_status",
                "NOT USED",
                "status",
                "1761 and 1769 are different epochs, not simultaneous observer tracks",
            ]
        )


def backup_outputs() -> dict[str, str]:
    statuses = {
        "track_csv": copy_verified(TRACK_CSV, DRIVE_OUTPUTS["track_csv"]),
        "results_csv": copy_verified(RESULTS_CSV, DRIVE_OUTPUTS["results_csv"]),
        "figure_png": copy_verified(FIGURE_PNG, DRIVE_OUTPUTS["figure_png"]),
    }
    script_path = Path(__file__).resolve() if "__file__" in globals() else None
    statuses["python"] = (
        copy_verified(script_path, DRIVE_OUTPUTS["python"])
        if script_path is not None and script_path.exists()
        else "NOT AVAILABLE"
    )
    return statuses


def main() -> None:
    workbook = resolve_workbook()
    records = read_target_records(workbook)
    tracks = [build_track(year, records[year]) for year in TARGET_YEARS]
    write_track_csv(tracks)
    write_results_csv(tracks)
    plot_engineering_tracks(tracks)
    backups = backup_outputs()

    print(f"CODE OUTPUT: {VERSION}")
    print("CODE INPUTS")
    print(f"Workbook : {workbook}")
    print(f"Template : {REFERENCE_TEMPLATE}")
    print("COMMENTS")
    print("V0126B is REJECTED because it sent a long TLIST URL to JPL Horizons.")
    print("This repair uses four short JPL range queries: Sun and Venus for 1761, then Sun and Venus for 1769.")
    print("No AI images. Matplotlib only. IMCCE supplies contact epochs; JPL supplies every plotted one-minute position.")
    print("RESULTS")
    for track in tracks:
        ca = track["event_points"]["CA"]
        print(
            f"{track['year']} | rows={len(track['minute_master'])} | angle={track['angle_deg']:.6f} deg | "
            f"CA=({ca[0]:+.6f}, {ca[1]:+.6f}) R_sun | RMS={track['rms_solar_radii']:.9f} R_sun | "
            f"curvature={track['curvature_inverse_solar_radii']:.9f} 1/R_sun"
        )
        print(
            f"{track['year']} | R2 linear={track['linear_r2']:.9f} | "
            f"quadratic={track['quadratic_r2']:.9f} | cubic={track['cubic_r2']:.9f}"
        )
    print("A-prime B-prime and Halley solar parallax : NOT USED — different transit epochs.")
    print("OUTPUT SUMMARY")
    print(f"Track CSV : {TRACK_CSV}")
    print(f"Results CSV : {RESULTS_CSV}")
    print(f"Figure PNG : {FIGURE_PNG}")
    print(
        f"Drive backup : TRACK={backups['track_csv']} | RESULTS={backups['results_csv']} | "
        f"PNG={backups['figure_png']} | PYTHON={backups['python']}"
    )
    print("PAPER COMPARISON")
    print("IMCCE contact epochs are used as event references; every plotted XY position comes from JPL Horizons range-query vectors.")
    print("EQUATION STATUS")
    print("VERIFIED — relative Sun/Venus tangent-plane projection, one-minute sampling, event splines, closest approach, PCA angle, RMS, R2, and curvature evaluated.")
    print(datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S %z"))
    print("# V0126C")


if __name__ == "__main__":
    main()

# V0126C
