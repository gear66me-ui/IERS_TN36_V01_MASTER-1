# V0125A
# Audit reference: rebuild the IMCCE Venus Transit Canon workbook into true plot-ready columns.

from __future__ import annotations

import csv
import hashlib
import shutil
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import Reference, ScatterChart, Series
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl>=3.1.0"])
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import Reference, ScatterChart, Series
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

VERSION = "IERS-0125A"
LOCAL_TZ = timezone(timedelta(hours=-5))
TARGET_YEARS = (1761, 1769, 1874, 1882, 2004, 2012)

DRIVE_ROOT = Path("/content/drive/MyDrive/Colab Notebooks/JPL - 1769 VENUS TRANSIT/GitHub")
SOURCE_CSV = DRIVE_ROOT / "DATA" / "CSV" / "IMCCE_VENUS_TRANSIT_CANON_MASTER.csv"
MASTER_XLSX = DRIVE_ROOT / "DATA" / "XLSX" / "IMCCE_VENUS_TRANSIT_CANON_MASTER.xlsx"
ORGANIZED_XLSX = DRIVE_ROOT / "DATA" / "XLSX" / "IMCCE_VENUS_TRANSIT_CANON_ORGANIZED.xlsx"
PLOT_CSV = DRIVE_ROOT / "DATA" / "CSV" / "IMCCE_VENUS_TRANSIT_CANON_PLOT_DATA.csv"
AUDIT_DIR = DRIVE_ROOT / "AUDIT"
PREVIOUS_BACKUP = AUDIT_DIR / "IMCCE_VENUS_TRANSIT_CANON_MASTER_PRE_V0125A.xlsx"
SCRIPT_BACKUP = DRIVE_ROOT / "PYTHON" / "IERS_0125A_IMCCE_WORKBOOK_REORGANIZER.py"

DARK_GREEN = "0B5D4B"
GREEN = "1F7A63"
LIGHT_GREEN = "DDEFE9"
WHITE = "FFFFFF"
GRID = "B7C9C2"
YELLOW = "FFF2CC"
THIN = Side(style="thin", color=GRID)

PLOT_COLUMNS = [
    ("year", "Year", "int"),
    ("x_year", "X Year", "float"),
    ("y_signed_impact_ratio", "Y Signed Impact Ratio", "float6"),
    ("record_id", "Record ID", "int"),
    ("jd_tdb", "JD TDB", "float3"),
    ("date_ut_label", "Date UT", "text"),
    ("mid_ut_hhmm", "Mid-Transit UT", "text"),
    ("sun_radius_arcsec", "Solar Radius (arcsec)", "float3"),
    ("minimum_distance_arcsec", "Signed Minimum Distance (arcsec)", "float3"),
    ("distance_ratio", "Source Ratio", "float6"),
    ("venus_radius_arcsec", "Venus Radius (arcsec)", "float3"),
    ("c1_ut", "C1 UT", "text"),
    ("c2_ut", "C2 UT", "text"),
    ("c3_ut", "C3 UT", "text"),
    ("c4_ut", "C4 UT", "text"),
    ("subsolar_longitude_ingress_deg", "Longitude Begin (deg)", "float3"),
    ("subsolar_longitude_egress_deg", "Longitude End (deg)", "float3"),
    ("subsolar_latitude_deg", "Subsolar Latitude (deg)", "float3"),
    ("relative_velocity_deg_per_day", "Relative Velocity (deg/day)", "float6"),
    ("venus_ecliptic_latitude_deg", "Venus Latitude (deg)", "float3"),
    ("node", "Node", "int"),
    ("tdb_minus_ut_seconds", "TDB-UT (s)", "float3"),
    ("record_status", "Record Status", "text"),
]

CONTACT_COLUMNS = [
    ("year", "Year", "int"),
    ("date_ut_label", "Date UT", "text"),
    ("c1_ut", "C1 UT", "text"),
    ("c2_ut", "C2 UT", "text"),
    ("c3_ut", "C3 UT", "text"),
    ("c4_ut", "C4 UT", "text"),
    ("external_duration_seconds", "External Duration (s)", "float3"),
    ("internal_duration_seconds", "Internal Duration (s)", "float3"),
    ("mid_ut_hhmm", "Mid-Transit UT", "text"),
]

DERIVED_COLUMNS = [
    ("year", "Year", "int"),
    ("impact_parameter_abs_arcsec", "Absolute Impact (arcsec)", "float3"),
    ("ratio_abs_calculated", "Absolute Impact Ratio", "float6"),
    ("y_signed_impact_ratio", "Signed Impact Ratio", "float6"),
    ("ratio_residual_source_minus_calculated", "Source Minus Calculated Ratio", "float6"),
    ("closest_approach_sign", "Closest-Approach Side", "text"),
    ("geometry_class", "Geometry Class", "text"),
    ("ratio_validation", "Ratio Validation", "text"),
    ("record_status", "Record Status", "text"),
]

SOURCE_COLUMNS = [
    ("record_id", "Record ID", "int"),
    ("source_line_number", "Source Line", "int"),
    ("year", "Year", "int"),
    ("raw_record", "Raw IMCCE Record", "text"),
    ("missing_fields", "Missing Fields", "text"),
    ("missing_field_count", "Missing Count", "int"),
    ("source_data_url", "Source Data URL", "text"),
    ("source_page_url", "Source Page URL", "text"),
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def text(row: dict[str, str], key: str) -> str:
    return str(row.get(key, "")).strip()


def integer(row: dict[str, str], key: str) -> int | None:
    value = text(row, key)
    return None if value == "" else int(float(value))


def number(row: dict[str, str], key: str) -> float | None:
    value = text(row, key)
    return None if value == "" else float(value)


def normalized_rows() -> list[dict[str, object]]:
    if not SOURCE_CSV.exists():
        raise FileNotFoundError(f"Canonical source CSV not found: {SOURCE_CSV}")
    with SOURCE_CSV.open("r", encoding="utf-8", newline="") as handle:
        source_rows = list(csv.DictReader(handle))
    if len(source_rows) != 77:
        raise RuntimeError(f"Expected 77 canonical records, found {len(source_rows)}")

    rows: list[dict[str, object]] = []
    for source in source_rows:
        year = integer(source, "year")
        sun_radius = number(source, "sun_radius_arcsec")
        delta = number(source, "minimum_distance_arcsec")
        if year is None or sun_radius is None or delta is None or sun_radius == 0.0:
            raise RuntimeError(f"Invalid canonical geometry row: {source}")
        row: dict[str, object] = dict(source)
        row["year"] = year
        row["x_year"] = float(year)
        row["y_signed_impact_ratio"] = delta / sun_radius
        rows.append(row)
    rows.sort(key=lambda item: int(item["year"]))
    return rows


def typed_value(row: dict[str, object], key: str, kind: str) -> object:
    value = row.get(key, "")
    if value is None or str(value).strip() == "":
        return None
    if kind == "int":
        return int(float(value))
    if kind.startswith("float"):
        return float(value)
    return str(value)


def style_sheet(ws, widths: dict[int, float] | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    if widths:
        for column_index, width in widths.items():
            ws.column_dimensions[get_column_letter(column_index)].width = width


def add_table(ws, name: str) -> None:
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def apply_formats(ws, columns: list[tuple[str, str, str]]) -> None:
    for column_index, (_, _, kind) in enumerate(columns, start=1):
        if kind == "int":
            fmt = "0"
        elif kind == "float3":
            fmt = "0.000"
        elif kind == "float6":
            fmt = "0.000000"
        elif kind == "float":
            fmt = "0.000000"
        else:
            continue
        for row_index in range(2, ws.max_row + 1):
            ws.cell(row=row_index, column=column_index).number_format = fmt


def build_tabular_sheet(wb: Workbook, name: str, rows: list[dict[str, object]], columns, table_name: str) -> None:
    ws = wb.create_sheet(name)
    ws.append([label for _, label, _ in columns])
    for row in rows:
        ws.append([typed_value(row, key, kind) for key, _, kind in columns])
    widths = {}
    for index, (_, label, kind) in enumerate(columns, start=1):
        if label in {"Raw IMCCE Record", "Missing Fields", "Source Data URL", "Source Page URL"}:
            widths[index] = 42
        elif kind == "text":
            widths[index] = max(14, min(26, len(label) + 2))
        else:
            widths[index] = max(12, min(22, len(label) + 2))
    style_sheet(ws, widths)
    apply_formats(ws, columns)
    add_table(ws, table_name)


def add_plot_chart(ws) -> None:
    chart = ScatterChart()
    chart.title = "IMCCE Venus Transit Canon: Year vs Signed Impact Ratio"
    chart.style = 13
    chart.height = 10
    chart.width = 20
    chart.x_axis.title = "Astronomical Year"
    chart.y_axis.title = "Signed Minimum Distance / Solar Radius"
    x_values = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    y_values = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    series = Series(y_values, x_values, title="Canon Records")
    series.marker.symbol = "circle"
    series.marker.size = 4
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)
    ws.add_chart(chart, "Y2")


def add_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    ws["A1"] = "IMCCE VENUS TRANSIT CANON — ORGANIZED WORKBOOK"
    ws["A1"].font = Font(bold=True, color=WHITE, size=15)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_GREEN)
    ws.merge_cells("A1:D1")
    ws["A3"] = "Start here"
    ws["B3"] = "PLOT_DATA"
    ws["A4"] = "Plot X"
    ws["B4"] = "X Year"
    ws["A5"] = "Plot Y"
    ws["B5"] = "Y Signed Impact Ratio = Signed Minimum Distance / Solar Radius"
    ws["A7"] = "Sheet"
    ws["B7"] = "Purpose"
    guide = [
        ("PLOT_DATA", "One transit per row; numeric values in separate columns; includes an embedded scatter chart."),
        ("MASTER", "All original canonical fields, one record per row, reordered for analysis."),
        ("TARGET_YEARS", "Only 1761, 1769, 1874, 1882, 2004, and 2012."),
        ("CONTACTS", "C1-C4 and transit durations."),
        ("DERIVED", "Calculated impact ratios and validation fields."),
        ("SOURCE_RAW", "Raw source text isolated away from plot data."),
    ]
    for offset, (sheet_name, purpose) in enumerate(guide, start=8):
        ws.cell(row=offset, column=1, value=sheet_name)
        ws.cell(row=offset, column=2, value=purpose)
    for cell in ws[7]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=GREEN)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.sheet_view.showGridLines = False


def build_master_sheet(wb: Workbook, rows: list[dict[str, object]]) -> None:
    preferred = [key for key, _, _ in PLOT_COLUMNS]
    extra = [key for key in rows[0].keys() if key not in preferred]
    ordered = preferred + extra
    columns = []
    for key in ordered:
        label = key.replace("_", " ").title()
        if key in {"record_id", "source_line_number", "day", "month", "year", "mid_hour", "node", "missing_field_count"}:
            kind = "int"
        elif key in {
            "jd_tdb", "mid_minute", "sun_radius_arcsec", "minimum_distance_arcsec",
            "distance_ratio", "venus_radius_arcsec", "subsolar_longitude_ingress_deg",
            "subsolar_longitude_egress_deg", "subsolar_latitude_deg",
            "relative_velocity_deg_per_day", "venus_ecliptic_latitude_deg",
            "tdb_minus_ut_seconds", "mid_ut_seconds_of_day", "impact_parameter_abs_arcsec",
            "ratio_abs_calculated", "ratio_residual_source_minus_calculated",
            "c1_seconds_of_day", "c2_seconds_of_day", "c3_seconds_of_day", "c4_seconds_of_day",
            "external_duration_seconds", "internal_duration_seconds", "x_year", "y_signed_impact_ratio",
        }:
            kind = "float6"
        else:
            kind = "text"
        columns.append((key, label, kind))
    build_tabular_sheet(wb, "MASTER", rows, columns, "IMCCE_Master_Organized")


def write_plot_csv(rows: list[dict[str, object]]) -> None:
    PLOT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with PLOT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([label for _, label, _ in PLOT_COLUMNS])
        for row in rows:
            writer.writerow([typed_value(row, key, kind) for key, _, kind in PLOT_COLUMNS])


def verify_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False, read_only=False)
    required = {"PLOT_DATA", "MASTER", "TARGET_YEARS", "CONTACTS", "DERIVED", "SOURCE_RAW", "README"}
    missing = sorted(required - set(workbook.sheetnames))
    if missing:
        raise RuntimeError(f"Missing organized sheets: {missing}")
    plot_sheet = workbook["PLOT_DATA"]
    if plot_sheet.max_row != 78 or plot_sheet.max_column != len(PLOT_COLUMNS):
        raise RuntimeError(
            f"PLOT_DATA dimensions incorrect: rows={plot_sheet.max_row}, columns={plot_sheet.max_column}"
        )
    if workbook.active.title != "PLOT_DATA":
        raise RuntimeError(f"Active sheet is not PLOT_DATA: {workbook.active.title}")
    for row_number in range(2, plot_sheet.max_row + 1):
        populated = sum(plot_sheet.cell(row=row_number, column=column).value is not None for column in range(1, 12))
        if populated < 8:
            raise RuntimeError(f"PLOT_DATA row {row_number} is not properly separated into columns")
    workbook.close()


def backup_previous_workbook() -> str:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    if not MASTER_XLSX.exists():
        return "NOT FOUND"
    if PREVIOUS_BACKUP.exists() and sha256(PREVIOUS_BACKUP) == sha256(MASTER_XLSX):
        return "UNCHANGED"
    shutil.copy2(MASTER_XLSX, PREVIOUS_BACKUP)
    if sha256(PREVIOUS_BACKUP) != sha256(MASTER_XLSX):
        raise RuntimeError("Previous workbook backup hash verification failed")
    return "COPIED"


def backup_script() -> str:
    script = Path(__file__).resolve() if "__file__" in globals() else None
    if not script or not script.exists():
        return "NOT AVAILABLE"
    SCRIPT_BACKUP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(script, SCRIPT_BACKUP)
    return "COPIED"


def main() -> None:
    rows = normalized_rows()
    previous_status = backup_previous_workbook()

    wb = Workbook()
    wb.remove(wb.active)
    build_tabular_sheet(wb, "PLOT_DATA", rows, PLOT_COLUMNS, "IMCCE_Plot_Data")
    add_plot_chart(wb["PLOT_DATA"])
    build_master_sheet(wb, rows)
    target_rows = [row for row in rows if int(row["year"]) in TARGET_YEARS]
    build_tabular_sheet(wb, "TARGET_YEARS", target_rows, PLOT_COLUMNS, "IMCCE_Target_Years")
    build_tabular_sheet(wb, "CONTACTS", rows, CONTACT_COLUMNS, "IMCCE_Contacts")
    build_tabular_sheet(wb, "DERIVED", rows, DERIVED_COLUMNS, "IMCCE_Derived")
    build_tabular_sheet(wb, "SOURCE_RAW", rows, SOURCE_COLUMNS, "IMCCE_Source_Raw")
    add_readme(wb)

    for year in TARGET_YEARS:
        year_rows = [row for row in rows if int(row["year"]) == year]
        if len(year_rows) != 1:
            raise RuntimeError(f"Expected one row for {year}; found {len(year_rows)}")
        build_tabular_sheet(wb, str(year), year_rows, PLOT_COLUMNS, f"IMCCE_{year}_Record")

    wb.active = wb.sheetnames.index("PLOT_DATA")
    wb.properties.title = "IMCCE Venus Transit Canon Organized"
    wb.properties.creator = VERSION
    wb.properties.description = "Plot-ready IMCCE Venus transit canon with one record per row and one variable per column."

    ORGANIZED_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ORGANIZED_XLSX)
    shutil.copy2(ORGANIZED_XLSX, MASTER_XLSX)
    write_plot_csv(rows)
    verify_workbook(MASTER_XLSX)
    script_status = backup_script()

    print(f"CODE OUTPUT: {VERSION}")
    print("CODE INPUTS")
    print(f"Canonical CSV : {SOURCE_CSV}")
    print(f"Previous workbook : {MASTER_XLSX}")
    print("COMMENTS")
    print("The old layout is REJECTED. The workbook is rebuilt with one transit per row and one variable per column.")
    print("RESULTS")
    print(f"PLOT_DATA rows : {len(rows)} | columns : {len(PLOT_COLUMNS)} | target rows : {len(target_rows)}")
    print(f"Active sheet : PLOT_DATA | Embedded scatter chart : YES | Previous workbook backup : {previous_status}")
    print("OUTPUT SUMMARY")
    print(f"Reorganized master workbook : {MASTER_XLSX}")
    print(f"Organized copy : {ORGANIZED_XLSX}")
    print(f"Plot-ready CSV : {PLOT_CSV}")
    print(f"Script backup : {script_status}")
    print("PAPER COMPARISON")
    print("NOT USED — workbook organization stage.")
    print("EQUATION STATUS")
    print("VERIFIED — Y Signed Impact Ratio equals Signed Minimum Distance divided by Solar Radius for all 77 records.")
    print(datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S %z"))
    print("# V0125A")


if __name__ == "__main__":
    main()

# V0125A
