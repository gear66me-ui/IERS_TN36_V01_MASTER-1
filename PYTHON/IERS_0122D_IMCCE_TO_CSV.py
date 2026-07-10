# V0122D
# Audit reference: Colab-compatible IMCCE Venus Transit Canon workbook builder using openpyxl.

from __future__ import annotations

import csv
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl>=3.1.0"])
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

VERSION = "IERS-0122D"
REVISION = "R2"
OUTPUT_ROOT = Path("/content/IERS_TN36_V01_MASTER_OUTPUT/IMCCE_VENUS_CANON")
MASTER_CSV = OUTPUT_ROOT / "IMCCE_VENUS_TRANSIT_CANON_MASTER.csv"
MASTER_XLSX = OUTPUT_ROOT / "IMCCE_VENUS_TRANSIT_CANON_MASTER.xlsx"
LOCAL_TZ = timezone(timedelta(hours=-5))
TARGET_YEARS = (1761, 1769, 1874, 1882, 2004, 2012)
SHEET_ORDER = ["README", "MASTER", "1761", "1769", "1874", "1882", "2004", "2012"]
SOURCE_PAGE_URL = "https://www.oca.eu/Mignard/Transits/Html/canon_venus.htm"
SOURCE_DATA_URL = "https://www.oca.eu/Mignard/Transits/Data/transit_venus.txt"

INT_COLUMNS = {
    "source_line_number", "day", "month", "year", "mid_hour", "node",
    "missing_field_count", "record_id",
}
FLOAT_COLUMNS = {
    "jd_tdb", "mid_minute", "sun_radius_arcsec", "minimum_distance_arcsec",
    "distance_ratio", "venus_radius_arcsec", "subsolar_longitude_ingress_deg",
    "subsolar_longitude_egress_deg", "subsolar_latitude_deg",
    "relative_velocity_deg_per_day", "venus_ecliptic_latitude_deg",
    "tdb_minus_ut_seconds", "mid_ut_seconds_of_day", "impact_parameter_abs_arcsec",
    "ratio_abs_calculated", "ratio_residual_source_minus_calculated",
    "c1_seconds_of_day", "c2_seconds_of_day", "c3_seconds_of_day", "c4_seconds_of_day",
    "external_duration_seconds", "internal_duration_seconds",
}
WIDE_COLUMNS = {
    "raw_record": 46,
    "missing_fields": 38,
    "source_page_url": 42,
    "source_data_url": 42,
    "geometry_class": 30,
    "record_status": 22,
    "node_label": 18,
    "year_display": 18,
}

DARK_GREEN = "0B5D4B"
GREEN = "1F7A63"
LIGHT_GREEN = "DDEFE9"
PALE_GREEN = "EAF5F1"
WHITE = "FFFFFF"
DARK_TEXT = "163D33"
YELLOW = "FFF2CC"
RED = "F4CCCC"
GRID = "B7C9C2"
THIN = Side(style="thin", color=GRID)


def typed_value(header: str, value: str) -> object:
    text = str(value).strip()
    if text == "":
        return None
    if header in INT_COLUMNS:
        return int(float(text))
    if header in FLOAT_COLUMNS:
        return float(text)
    return text


def load_master() -> tuple[list[str], list[list[object]]]:
    if not MASTER_CSV.exists():
        raise FileNotFoundError(f"Run IERS-0122C first: {MASTER_CSV}")
    with MASTER_CSV.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        rows = [[typed_value(header, record.get(header, "")) for header in headers] for record in reader]
    if len(rows) != 77:
        raise RuntimeError(f"Expected 77 master rows, found {len(rows)}")
    return headers, rows


def style_title(ws, last_column: int, subtitle: str) -> None:
    end = get_column_letter(last_column)
    ws.merge_cells(f"A1:{end}1")
    ws["A1"] = "IMCCE VENUS TRANSIT CANON"
    ws["A1"].font = Font(bold=True, color=WHITE, size=16)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_GREEN)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{end}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color=DARK_TEXT, size=10)
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")


def set_column_widths(ws, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        width = WIDE_COLUMNS.get(header, 14)
        if header.endswith("_ut") or header in {"date_ut_label", "mid_ut_hhmm"}:
            width = 16
        ws.column_dimensions[get_column_letter(index)].width = width


def apply_number_formats(ws, headers: list[str], first_row: int, last_row: int) -> None:
    for index, header in enumerate(headers, start=1):
        number_format = None
        if header in INT_COLUMNS:
            number_format = "0"
        elif header == "jd_tdb":
            number_format = "0.000"
        elif "ratio" in header:
            number_format = "0.000000"
        elif header in FLOAT_COLUMNS:
            number_format = "0.000"
        if number_format:
            for row in range(first_row, last_row + 1):
                ws.cell(row=row, column=index).number_format = number_format


def add_table_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = wb.create_sheet(name)
    subtitle = "All 77 canon records" if name == "MASTER" else f"Canonical record for the {name} Venus transit"
    style_title(ws, len(headers), subtitle)
    ws.append([])
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_row = 4
    first_data_row = 5
    last_row = 4 + len(rows)
    last_column = len(headers)
    end = get_column_letter(last_column)

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color=WHITE, size=9)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.row_dimensions[header_row].height = 34

    for row in ws.iter_rows(min_row=first_data_row, max_row=last_row, min_col=1, max_col=last_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column <= last_column)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    table = Table(displayName=f"IMCCE_{name}_Table", ref=f"A4:{end}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, headers)
    apply_number_formats(ws, headers, first_data_row, last_row)

    status_column = headers.index("record_status") + 1
    status_letter = get_column_letter(status_column)
    status_fill = PatternFill("solid", fgColor=YELLOW)
    ws.conditional_formatting.add(
        f"A{first_data_row}:{end}{last_row}",
        FormulaRule(formula=[f'${status_letter}{first_data_row}="SOURCE_HEADER_ONLY"'], fill=status_fill),
    )

    ratio_column = headers.index("ratio_validation") + 1
    ratio_letter = get_column_letter(ratio_column)
    ratio_fill = PatternFill("solid", fgColor=RED)
    ws.conditional_formatting.add(
        f"{ratio_letter}{first_data_row}:{ratio_letter}{last_row}",
        FormulaRule(formula=[f'{ratio_letter}{first_data_row}="REVIEW"'], fill=ratio_fill),
    )


def add_readme(wb: Workbook, row_count: int, complete_count: int, header_only_count: int) -> None:
    ws = wb.create_sheet("README")
    ws.merge_cells("A1:D1")
    ws["A1"] = "IMCCE VENUS TRANSIT CANON — WORKBOOK README"
    ws["A1"].font = Font(bold=True, color=WHITE, size=16)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_GREEN)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    content = [
        ["Workbook", MASTER_XLSX.name, "Version", f"{VERSION} {REVISION}"],
        ["Source page", SOURCE_PAGE_URL, "Source data", SOURCE_DATA_URL],
        ["Master records", row_count, "Complete records", complete_count],
        ["Header-only records", header_only_count, "Target-year sheets", len(TARGET_YEARS)],
        [None, None, None, None],
        ["SHEET GUIDE", None, None, None],
        ["MASTER", "Normalized 77-record canon with source and calculated fields.", None, None],
        ["1761–2012", "One filtered record per requested historical transit year.", None, None],
        [None, None, None, None],
        ["TRACEABILITY", None, None, None],
        ["Source fields", "Preserved from the IMCCE text canon.", None, None],
        ["Calculated fields", "Appended by IERS-0122C and explicitly named.", None, None],
        ["SOURCE_HEADER_ONLY", "The IMCCE source omitted contact and trailing geometry fields.", None, None],
        ["REVIEW", "Validation flag requiring inspection; no ratio reviews existed at build time.", None, None],
    ]
    for row_index, row in enumerate(content, start=3):
        for column_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=column_index, value=value)

    for section_row in (8, 12):
        for cell in ws[section_row]:
            cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
            cell.font = Font(bold=True, color=DARK_TEXT)
    for row in ws.iter_rows(min_row=3, max_row=16, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 46
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws["B4"].hyperlink = SOURCE_PAGE_URL
    ws["D4"].hyperlink = SOURCE_DATA_URL
    ws["B4"].style = "Hyperlink"
    ws["D4"].style = "Hyperlink"


def verify_workbook(path: Path) -> None:
    check = load_workbook(path, data_only=False, read_only=False)
    if check.sheetnames != SHEET_ORDER:
        raise RuntimeError(f"Workbook sheet order mismatch: {check.sheetnames}")
    if check["MASTER"].max_row != 81:
        raise RuntimeError(f"MASTER row count mismatch: {check['MASTER'].max_row}")
    for year in TARGET_YEARS:
        sheet = check[str(year)]
        if sheet.max_row != 5:
            raise RuntimeError(f"Sheet {year} should contain one data row; max_row={sheet.max_row}")
    for ws in check.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#") and cell.value in {
                    "#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"
                }:
                    raise RuntimeError(f"Spreadsheet error in {ws.title}!{cell.coordinate}: {cell.value}")
    check.close()


def main() -> None:
    headers, rows = load_master()
    year_index = headers.index("year")
    status_index = headers.index("record_status")
    complete_count = sum(row[status_index] == "COMPLETE" for row in rows)
    header_only_count = sum(row[status_index] == "SOURCE_HEADER_ONLY" for row in rows)

    wb = Workbook()
    wb.remove(wb.active)
    add_readme(wb, len(rows), complete_count, header_only_count)
    add_table_sheet(wb, "MASTER", headers, rows)
    for year in TARGET_YEARS:
        selected = [row for row in rows if row[year_index] == year]
        if len(selected) != 1:
            raise RuntimeError(f"Expected exactly one record for {year}, found {len(selected)}")
        add_table_sheet(wb, str(year), headers, selected)

    wb.properties.title = "IMCCE Venus Transit Canon Master"
    wb.properties.subject = "IMCCE Venus transit canon parsed and normalized by IERS-0122"
    wb.properties.creator = "IERS-0122"
    wb.properties.description = "77 IMCCE Venus transit canon records with six dedicated historical transit sheets."
    MASTER_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(MASTER_XLSX)
    verify_workbook(MASTER_XLSX)

    print(f"CODE OUTPUT: {VERSION} {REVISION}")
    print("CODE INPUTS")
    print(f"Master CSV : {MASTER_CSV}")
    print("COMMENTS")
    print("Colab-compatible workbook created with openpyxl, styled tables, frozen headers, and traceability notes.")
    print("RESULTS")
    print(f"Rows : {len(rows)} | Sheets : {len(SHEET_ORDER)} | Complete : {complete_count} | Header-only : {header_only_count}")
    print("OUTPUT SUMMARY")
    print(f"Workbook : {MASTER_XLSX}")
    print("PAPER COMPARISON")
    print("NOT USED — workbook publication stage.")
    print("EQUATION STATUS")
    print("VERIFIED — sheet order, row counts, target-year sheets, and spreadsheet-error scan passed.")
    print(datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S %z"))
    print("# V0122D")


if __name__ == "__main__":
    main()

# V0122D
