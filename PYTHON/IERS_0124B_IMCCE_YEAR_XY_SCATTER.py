# V0124B
# Audit reference: extract actual IMCCE workbook values and plot year versus signed normalized closest approach.

from __future__ import annotations

import csv
import hashlib
import shutil
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl>=3.1.0"])
    from openpyxl import load_workbook

import matplotlib.pyplot as plt
import numpy as np

VERSION = "IERS-0124B"
LOCAL_TZ = timezone(timedelta(hours=-5))
TARGET_YEARS = (1761, 1769, 1874, 1882, 2004, 2012)

DRIVE_ROOT = Path("/content/drive/MyDrive/Colab Notebooks/JPL - 1769 VENUS TRANSIT/GitHub")
DRIVE_XLSX = DRIVE_ROOT / "DATA" / "XLSX" / "IMCCE_VENUS_TRANSIT_CANON_MASTER.xlsx"
LOCAL_XLSX = Path(
    "/content/IERS_TN36_V01_MASTER_OUTPUT/IMCCE_VENUS_CANON/"
    "IMCCE_VENUS_TRANSIT_CANON_MASTER.xlsx"
)
OUTPUT_ROOT = Path(
    "/content/IERS_TN36_V01_MASTER_OUTPUT/IMCCE_VENUS_CANON/PLOTS/V0124B"
)
XY_CSV = OUTPUT_ROOT / "IERS_0124B_IMCCE_YEAR_XY.csv"
RESULTS_CSV = OUTPUT_ROOT / "IERS_0124B_IMCCE_YEAR_XY_RESULTS.csv"
FIGURE_PNG = OUTPUT_ROOT / "IERS_0124B_IMCCE_YEAR_VS_SIGNED_IMPACT.png"
DRIVE_XY_CSV = DRIVE_ROOT / "DATA" / "CSV" / XY_CSV.name
DRIVE_RESULTS_CSV = DRIVE_ROOT / "DATA" / "CSV" / RESULTS_CSV.name
DRIVE_FIGURE = DRIVE_ROOT / "DATA" / "PNG" / FIGURE_PNG.name
DRIVE_SCRIPT = DRIVE_ROOT / "PYTHON" / "IERS_0124B_IMCCE_YEAR_XY_SCATTER.py"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def copy_verified(source: Path, destination: Path) -> str:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists() and sha256(destination) == sha256(source):
        return "UNCHANGED"
    shutil.copy2(source, destination)
    if sha256(source) != sha256(destination):
        raise RuntimeError(f"SHA-256 verification failed: {destination}")
    return "COPIED"


def resolve_workbook() -> Path:
    if DRIVE_XLSX.exists():
        return DRIVE_XLSX
    if LOCAL_XLSX.exists():
        return LOCAL_XLSX
    raise FileNotFoundError("IMCCE workbook not found. Run IERS-0122D first.")


def read_master_sheet(path: Path) -> list[dict[str, float]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["MASTER"]
    headers = [sheet.cell(row=4, column=column).value for column in range(1, sheet.max_column + 1)]
    index = {str(name): position + 1 for position, name in enumerate(headers) if name is not None}
    required = {"year", "minimum_distance_arcsec", "sun_radius_arcsec", "record_id"}
    missing = sorted(required - set(index))
    if missing:
        raise RuntimeError(f"MASTER sheet missing required columns: {missing}")

    rows: list[dict[str, float]] = []
    for row_number in range(5, sheet.max_row + 1):
        year_value = sheet.cell(row=row_number, column=index["year"]).value
        if year_value is None:
            continue
        year = int(year_value)
        delta = float(sheet.cell(row=row_number, column=index["minimum_distance_arcsec"]).value)
        sun_radius = float(sheet.cell(row=row_number, column=index["sun_radius_arcsec"]).value)
        record_id = int(sheet.cell(row=row_number, column=index["record_id"]).value)
        if sun_radius <= 0.0:
            raise RuntimeError(f"Invalid solar radius for year {year}: {sun_radius}")
        rows.append(
            {
                "year": year,
                "x": float(year),
                "y": delta / sun_radius,
                "delta_arcsec": delta,
                "sun_radius_arcsec": sun_radius,
                "record_id": record_id,
            }
        )
    workbook.close()
    rows.sort(key=lambda item: item["year"])
    if len(rows) != 77:
        raise RuntimeError(f"Expected 77 IMCCE rows, found {len(rows)}")
    if len({int(row['year']) for row in rows}) != len(rows):
        raise RuntimeError("Duplicate years detected in MASTER sheet")
    return rows


def save_xy(rows: list[dict[str, float]]) -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    with XY_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["year", "x", "y"])
        writer.writeheader()
        writer.writerows({"year": int(row["year"]), "x": row["x"], "y": row["y"]} for row in rows)


def save_results(rows: list[dict[str, float]]) -> None:
    y_values = np.asarray([row["y"] for row in rows], dtype=float)
    selected = {int(row["year"]): float(row["y"]) for row in rows if int(row["year"]) in TARGET_YEARS}
    with RESULTS_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["quantity", "value", "unit", "traceability"])
        writer.writerow(["record_count", len(rows), "records", "MASTER worksheet"])
        writer.writerow(["x_definition", "astronomical_year", "year", "MASTER.year"])
        writer.writerow(["y_definition", "minimum_distance_arcsec / sun_radius_arcsec", "dimensionless", "calculated"])
        writer.writerow(["minimum_y", float(y_values.min()), "dimensionless", "calculated"])
        writer.writerow(["maximum_y", float(y_values.max()), "dimensionless", "calculated"])
        writer.writerow(["mean_y", float(y_values.mean()), "dimensionless", "calculated"])
        writer.writerow(["rms_y", float(np.sqrt(np.mean(y_values**2))), "dimensionless", "calculated"])
        for year in TARGET_YEARS:
            writer.writerow([f"y_{year}", selected[year], "dimensionless", "MASTER worksheet"])
        writer.writerow(["halley_status", "NOT USED — no observer pair or topocentric tracks in this plot", "status", "project constraint"])


def plot_data(rows: list[dict[str, float]]) -> None:
    years = np.asarray([row["year"] for row in rows], dtype=float)
    y_values = np.asarray([row["y"] for row in rows], dtype=float)

    figure, axis = plt.subplots(figsize=(12.0, 6.8))
    axis.plot(years, y_values, linewidth=0.45, color="0.55", zorder=1)
    axis.scatter(years, y_values, s=9.0, color="black", zorder=2)
    axis.axhline(0.0, linewidth=0.45, color="0.45")
    axis.axhline(1.0, linewidth=0.35, color="0.65", linestyle="--")
    axis.axhline(-1.0, linewidth=0.35, color="0.65", linestyle="--")

    lookup = {int(row["year"]): row for row in rows}
    for year in TARGET_YEARS:
        point = lookup[year]
        axis.scatter([point["x"]], [point["y"]], s=24.0, facecolors="white", edgecolors="black", linewidths=0.65, zorder=3)
        offset = 0.055 if point["y"] <= 0.0 else -0.065
        axis.text(point["x"], point["y"] + offset, str(year), fontsize=7, ha="center", va="center")

    axis.set_xlabel("Astronomical year")
    axis.set_ylabel("Signed closest approach / solar radius")
    axis.set_title(
        "IMCCE Venus Transit Canon — Actual Workbook Data\n"
        "X = year, Y = signed minimum distance divided by solar radius"
    )
    axis.grid(False)
    axis.tick_params(width=0.5, labelsize=8)
    for spine in axis.spines.values():
        spine.set_linewidth(0.5)
    figure.tight_layout()
    figure.savefig(FIGURE_PNG, dpi=300, bbox_inches="tight")
    plt.show()
    plt.close(figure)


def backup_outputs() -> dict[str, str]:
    statuses = {
        "xy": copy_verified(XY_CSV, DRIVE_XY_CSV),
        "results": copy_verified(RESULTS_CSV, DRIVE_RESULTS_CSV),
        "figure": copy_verified(FIGURE_PNG, DRIVE_FIGURE),
    }
    script = Path(__file__).resolve() if "__file__" in globals() else None
    statuses["python"] = copy_verified(script, DRIVE_SCRIPT) if script and script.exists() else "NOT AVAILABLE"
    return statuses


def main() -> None:
    workbook_path = resolve_workbook()
    rows = read_master_sheet(workbook_path)
    save_xy(rows)
    save_results(rows)
    plot_data(rows)
    backup = backup_outputs()
    y_values = np.asarray([row["y"] for row in rows], dtype=float)

    print(f"CODE OUTPUT: {VERSION}")
    print("CODE INPUTS")
    print(f"Workbook : {workbook_path}")
    print("COMMENTS")
    print("V0124A is REJECTED. This version plots only values extracted directly from the MASTER worksheet.")
    print("RESULTS")
    print(f"Rows : {len(rows)} | X : astronomical year | Y : signed closest approach / solar radius")
    print(f"Y minimum : {y_values.min():.6f} | Y maximum : {y_values.max():.6f} | Y RMS : {np.sqrt(np.mean(y_values**2)):.6f}")
    print("OUTPUT SUMMARY")
    print(f"XY CSV : {XY_CSV}")
    print(f"Results CSV : {RESULTS_CSV}")
    print(f"Figure PNG : {FIGURE_PNG}")
    print(f"Drive backup : XY={backup['xy']} | RESULTS={backup['results']} | PNG={backup['figure']} | PYTHON={backup['python']}")
    print("PAPER COMPARISON")
    print("NOT USED — direct workbook-data visualization.")
    print("EQUATION STATUS")
    print("VERIFIED — Y equals signed minimum_distance_arcsec divided by sun_radius_arcsec for all 77 rows.")
    print(datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S %z"))
    print("# V0124B")


if __name__ == "__main__":
    main()

# V0124B
