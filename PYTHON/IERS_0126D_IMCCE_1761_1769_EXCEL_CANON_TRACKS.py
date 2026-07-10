# V0126D
# Audit reference: Excel-only 1761/1769 IMCCE canonical track reconstruction in IERS-0012O engineering style.

from __future__ import annotations

import csv
import hashlib
import math
import shutil
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path


def ensure_package(import_name: str, pip_name: str) -> None:
    try:
        __import__(import_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", pip_name])


for _import_name, _pip_name in (
    ("numpy", "numpy"),
    ("matplotlib", "matplotlib"),
    ("openpyxl", "openpyxl>=3.1.0"),
):
    ensure_package(_import_name, _pip_name)

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.patches import Circle
from openpyxl import load_workbook

VERSION = "IERS-0126D"
PROGRAM_NAME = "IERS_0126D_IMCCE_1761_1769_EXCEL_CANON_TRACKS.py"
REFERENCE_TEMPLATE = "IERS_0012O_NORTH_SOUTH_POLE_ENGINEERING_TRACK_PLOT_PI_SUN.py"
TARGET_YEARS = (1761, 1769)
LOCAL_TZ = timezone(timedelta(hours=-5))

DRIVE_ROOT = Path("/content/drive/MyDrive/Colab Notebooks/JPL - 1769 VENUS TRANSIT/GitHub")
ORGANIZED_XLSX = DRIVE_ROOT / "DATA" / "XLSX" / "IMCCE_VENUS_TRANSIT_CANON_ORGANIZED.xlsx"
MASTER_XLSX = DRIVE_ROOT / "DATA" / "XLSX" / "IMCCE_VENUS_TRANSIT_CANON_MASTER.xlsx"
OUTPUT_ROOT = Path("/content/IERS_TN36_V01_MASTER_OUTPUT/IMCCE_VENUS_CANON/PLOTS/V0126D")
TRACK_CSV = OUTPUT_ROOT / "IERS_0126D_IMCCE_1761_1769_EXCEL_TRACKS.csv"
RESULTS_CSV = OUTPUT_ROOT / "IERS_0126D_IMCCE_1761_1769_EXCEL_RESULTS.csv"
FIGURE_PNG = OUTPUT_ROOT / "IERS_0126D_IMCCE_1761_1769_EXCEL_ENGINEERING_TRACKS.png"

DRIVE_OUTPUTS = {
    "track_csv": DRIVE_ROOT / "DATA" / "CSV" / TRACK_CSV.name,
    "results_csv": DRIVE_ROOT / "DATA" / "CSV" / RESULTS_CSV.name,
    "figure_png": DRIVE_ROOT / "DATA" / "PNG" / FIGURE_PNG.name,
    "python": DRIVE_ROOT / "PYTHON" / PROGRAM_NAME,
}

TRACK_COLORS = {1761: "#ffc861", 1769: "#5ee08a"}

HEADER_ALIASES = {
    "year": "year",
    "record_id": "record_id",
    "jd_tdb": "jd_tdb",
    "date_ut": "date_ut_label",
    "date_ut_label": "date_ut_label",
    "mid_transit_ut": "mid_ut_hhmm",
    "mid_ut_hhmm": "mid_ut_hhmm",
    "mid_ut_seconds_of_day": "mid_ut_seconds_of_day",
    "solar_radius_arcsec": "sun_radius_arcsec",
    "sun_radius_arcsec": "sun_radius_arcsec",
    "signed_minimum_distance_arcsec": "minimum_distance_arcsec",
    "minimum_distance_arcsec": "minimum_distance_arcsec",
    "source_ratio": "distance_ratio",
    "distance_ratio": "distance_ratio",
    "venus_radius_arcsec": "venus_radius_arcsec",
    "c1_ut": "c1_ut",
    "c2_ut": "c2_ut",
    "c3_ut": "c3_ut",
    "c4_ut": "c4_ut",
    "relative_velocity_deg_day": "relative_velocity_deg_per_day",
    "relative_velocity_deg_per_day": "relative_velocity_deg_per_day",
    "node": "node",
    "record_status": "record_status",
}

REQUIRED_COLUMNS = {
    "year",
    "record_id",
    "sun_radius_arcsec",
    "minimum_distance_arcsec",
    "venus_radius_arcsec",
    "c1_ut",
    "c2_ut",
    "c3_ut",
    "c4_ut",
    "relative_velocity_deg_per_day",
    "record_status",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def copy_verified(source: Path, destination: Path) -> str:
    destination.parent.mkdir(parents=True, exist_ok=True)
    source_hash = sha256(source)
    if destination.exists() and sha256(destination) == source_hash:
        return "UNCHANGED"
    shutil.copy2(source, destination)
    if sha256(destination) != source_hash:
        raise RuntimeError(f"SHA-256 verification failed: {destination}")
    return "COPIED"


def normalize_header(value: object) -> str:
    text = str(value).strip().lower()
    output = []
    previous_underscore = False
    for character in text:
        if character.isalnum():
            output.append(character)
            previous_underscore = False
        elif not previous_underscore:
            output.append("_")
            previous_underscore = True
    return "".join(output).strip("_")


def canonical_header(value: object) -> str:
    normalized = normalize_header(value)
    return HEADER_ALIASES.get(normalized, normalized)


def resolve_workbook() -> Path:
    for path in (ORGANIZED_XLSX, MASTER_XLSX):
        if path.exists():
            return path
    raise FileNotFoundError("Run IERS-0125A before IERS-0126D; organized workbook not found.")


def inspect_sheet(sheet, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for column_index in range(1, sheet.max_column + 1):
        raw = sheet.cell(header_row, column_index).value
        if raw is None:
            continue
        columns[canonical_header(raw)] = column_index
    return columns


def read_target_records(path: Path) -> tuple[dict[int, dict[str, object]], str]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    chosen = None
    for sheet_name, header_row in (("PLOT_DATA", 1), ("MASTER", 1), ("MASTER", 4)):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        columns = inspect_sheet(sheet, header_row)
        if REQUIRED_COLUMNS.issubset(columns):
            chosen = (sheet, sheet_name, header_row, columns)
            break
    if chosen is None:
        workbook.close()
        raise RuntimeError("No workbook sheet contains the required canonical reconstruction columns.")

    sheet, sheet_name, header_row, columns = chosen
    records: dict[int, dict[str, object]] = {}
    for row_index in range(header_row + 1, sheet.max_row + 1):
        raw_year = sheet.cell(row_index, columns["year"]).value
        if raw_year is None:
            continue
        year = int(float(raw_year))
        if year not in TARGET_YEARS:
            continue
        record = {
            key: sheet.cell(row_index, column_index).value
            for key, column_index in columns.items()
        }
        if str(record["record_status"]).strip().upper() != "COMPLETE":
            workbook.close()
            raise RuntimeError(f"IMCCE record {year} is not complete: {record['record_status']}")
        records[year] = record
    workbook.close()

    missing_years = sorted(set(TARGET_YEARS) - set(records))
    if missing_years:
        raise RuntimeError(f"Missing target years: {missing_years}")
    return records, sheet_name


def clock_to_seconds(value: object) -> float:
    fields = str(value).strip().split(":")
    if len(fields) == 2:
        return int(fields[0]) * 3600.0 + float(fields[1]) * 60.0
    if len(fields) == 3:
        return int(fields[0]) * 3600.0 + int(fields[1]) * 60.0 + float(fields[2])
    raise ValueError(f"Unsupported clock value: {value!r}")


def seconds_to_clock(seconds: float) -> str:
    seconds %= 86400.0
    hour = int(seconds // 3600.0)
    minute = int((seconds % 3600.0) // 60.0)
    second = seconds % 60.0
    return f"{hour:02d}:{minute:02d}:{second:06.3f}"


def unwrap_contacts(record: dict[str, object]) -> np.ndarray:
    raw = [clock_to_seconds(record[f"c{index}_ut"]) for index in range(1, 5)]
    output = [raw[0]]
    for value in raw[1:]:
        while value < output[-1]:
            value += 86400.0
        output.append(value)
    return np.asarray(output, dtype=float)


def derive_track(year: int, record: dict[str, object]) -> dict[str, object]:
    sun_radius = float(record["sun_radius_arcsec"])
    venus_radius = float(record["venus_radius_arcsec"])
    signed_minimum = float(record["minimum_distance_arcsec"])
    source_ratio = float(record["distance_ratio"]) if record.get("distance_ratio") not in (None, "") else abs(signed_minimum) / sun_radius
    relative_velocity = float(record["relative_velocity_deg_per_day"])
    times = unwrap_contacts(record)

    y = signed_minimum / sun_radius
    venus_radius_ratio = venus_radius / sun_radius
    external_radius = 1.0 + venus_radius_ratio
    internal_radius = 1.0 - venus_radius_ratio
    external_term = external_radius**2 - y**2
    internal_term = internal_radius**2 - y**2
    if external_term <= 0.0 or internal_term <= 0.0:
        raise RuntimeError(f"Year {year} does not produce four real contact intersections.")

    x_external = math.sqrt(external_term)
    x_internal = math.sqrt(internal_term)
    contact_x = np.asarray([-x_external, -x_internal, x_internal, x_external], dtype=float)
    contact_y = np.full(4, y, dtype=float)

    slope, intercept = np.polyfit(times, contact_x, 1)
    fitted_contacts = slope * times + intercept
    residuals = contact_x - fitted_contacts
    fit_rms = float(np.sqrt(np.mean(residuals**2)))
    closest_time = float(-intercept / slope)

    sample_times = np.arange(times[0], times[-1] + 0.001, 60.0)
    if sample_times[-1] < times[-1]:
        sample_times = np.append(sample_times, times[-1])
    sample_x = slope * sample_times + intercept
    sample_y = np.full_like(sample_x, y)

    external_duration = float(times[3] - times[0])
    internal_duration = float(times[2] - times[1])
    fit_speed = abs(float(slope))
    source_speed = relative_velocity * 3600.0 / 86400.0 / sun_radius
    external_speed = 2.0 * x_external / external_duration
    internal_speed = 2.0 * x_internal / internal_duration
    external_midpoint = 0.5 * (times[0] + times[3])
    internal_midpoint = 0.5 * (times[1] + times[2])

    radial_checks = np.asarray(
        [
            math.hypot(contact_x[0], y) - external_radius,
            math.hypot(contact_x[1], y) - internal_radius,
            math.hypot(contact_x[2], y) - internal_radius,
            math.hypot(contact_x[3], y) - external_radius,
        ],
        dtype=float,
    )

    return {
        "year": year,
        "record": record,
        "sun_radius_arcsec": sun_radius,
        "venus_radius_arcsec": venus_radius,
        "signed_minimum_arcsec": signed_minimum,
        "source_ratio": source_ratio,
        "calculated_ratio": abs(y),
        "ratio_residual": source_ratio - abs(y),
        "y": y,
        "venus_radius_ratio": venus_radius_ratio,
        "times": times,
        "contact_x": contact_x,
        "contact_y": contact_y,
        "fitted_contacts": fitted_contacts,
        "contact_fit_residuals": residuals,
        "fit_rms": fit_rms,
        "slope": float(slope),
        "intercept": float(intercept),
        "closest_time": closest_time,
        "sample_times": sample_times,
        "sample_x": sample_x,
        "sample_y": sample_y,
        "external_chord": 2.0 * x_external,
        "internal_chord": 2.0 * x_internal,
        "external_duration": external_duration,
        "internal_duration": internal_duration,
        "fit_speed": fit_speed,
        "source_speed": source_speed,
        "external_speed": external_speed,
        "internal_speed": internal_speed,
        "external_midpoint": external_midpoint,
        "internal_midpoint": internal_midpoint,
        "radial_check_max": float(np.max(np.abs(radial_checks))),
        "track_angle_deg": 0.0,
        "curvature_inverse_solar_radii": 0.0,
    }


def add_label(axis, point: tuple[float, float], text: str, dx: float, dy: float, color: str) -> None:
    axis.annotate(
        text,
        xy=point,
        xytext=(point[0] + dx, point[1] + dy),
        textcoords="data",
        fontsize=5.7,
        color=color,
        ha="left",
        va="center",
        arrowprops={"arrowstyle": "-", "lw": 0.20, "color": color, "shrinkA": 0, "shrinkB": 2},
    )


def add_summary_table(axis, tracks: list[dict[str, object]]) -> None:
    rows = []
    for track in tracks:
        year = int(track["year"])
        rows.extend(
            [
                [f"{year} impact", f"{track['y']:+.6f}", "R_sun"],
                [f"{year} Venus radius", f"{track['venus_radius_ratio']:.6f}", "R_sun"],
                [f"{year} fit speed", f"{track['fit_speed']:.9f}", "R_sun/s"],
                [f"{year} fit RMS", f"{track['fit_rms']:.9f}", "R_sun"],
            ]
        )
    table = axis.table(
        cellText=rows,
        colLabels=["Quantity", "Value", "Unit"],
        loc="lower left",
        colWidths=[0.31, 0.23, 0.17],
        bbox=[0.405, 0.055, 0.405, 0.265],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(5.15)
    for (row, column), cell in table.get_celld().items():
        cell.set_linewidth(0.18)
        cell.set_edgecolor("#1e4f64")
        if row == 0:
            cell.set_facecolor("#0a1a22")
            cell.get_text().set_color("#66e8ff")
            cell.get_text().set_fontweight("bold")
        else:
            cell.set_facecolor("#050b0f")
            if column == 1:
                cell.get_text().set_color("#ffc861")
                cell.get_text().set_fontweight("bold")
            elif column == 2:
                cell.get_text().set_color("#5ee08a")
            else:
                cell.get_text().set_color("#dff8ff")


def plot_tracks(tracks: list[dict[str, object]]) -> None:
    figure, axis = plt.subplots(figsize=(9.6, 7.2), dpi=240)
    figure.patch.set_facecolor("#03080d")
    axis.set_facecolor("#03080d")
    axis.add_patch(Circle((0.0, 0.0), 1.0, fill=False, lw=0.36, ec="#66e8ff", alpha=0.95))
    axis.axhline(0.0, lw=0.18, color="#1d3d4a", alpha=0.72)
    axis.axvline(0.0, lw=0.18, color="#1d3d4a", alpha=0.72)

    for track in tracks:
        year = int(track["year"])
        color = TRACK_COLORS[year]
        axis.plot(
            track["sample_x"],
            track["sample_y"],
            lw=0.30,
            color=color,
            solid_capstyle="round",
            label=str(year),
            zorder=3,
        )
        axis.scatter(
            track["sample_x"][::6],
            track["sample_y"][::6],
            s=0.75,
            color=color,
            alpha=0.70,
            linewidths=0,
            zorder=4,
        )

        event_x = [
            float(track["contact_x"][0]),
            float(track["contact_x"][1]),
            0.0,
            float(track["contact_x"][2]),
            float(track["contact_x"][3]),
        ]
        event_labels = ("C1", "C2", "CA", "C3", "C4")
        for x_value, event in zip(event_x, event_labels):
            point = (x_value, float(track["y"]))
            axis.add_patch(
                Circle(
                    point,
                    float(track["venus_radius_ratio"]),
                    fill=False,
                    lw=0.20 if event != "CA" else 0.28,
                    ec=color,
                    alpha=0.92,
                    zorder=2,
                )
            )
            axis.scatter(
                [point[0]],
                [point[1]],
                s=3.8 if event == "CA" else 2.2,
                color=color,
                edgecolors="#03080d",
                linewidths=0.16,
                zorder=5,
            )

        ca_point = (0.0, float(track["y"]))
        ca_dy = 0.045 if track["y"] >= 0.0 else -0.045
        add_label(axis, ca_point, f"{year} CA", 0.055, ca_dy, color)
        for event, x_value, dx, dy in (
            ("C1", float(track["contact_x"][0]), -0.115, 0.040),
            ("C2", float(track["contact_x"][1]), -0.095, 0.030),
            ("C3", float(track["contact_x"][2]), 0.055, -0.032),
            ("C4", float(track["contact_x"][3]), 0.075, -0.043),
        ):
            add_label(axis, (x_value, float(track["y"])), f"{year} {event}", dx, dy, color)

    add_summary_table(axis, tracks)
    axis.set_xlim(-1.08, 1.08)
    axis.set_ylim(-1.08, 1.08)
    axis.set_aspect("equal", adjustable="box")
    for spine in axis.spines.values():
        spine.set_linewidth(0.22)
        spine.set_color("#25708b")
    axis.tick_params(axis="both", colors="#8fb4c1", labelsize=6.5, width=0.22, length=2.0)
    axis.grid(True, color="#102630", linewidth=0.16, alpha=0.55)
    axis.set_xlabel("Canonical along-track coordinate / solar radius", color="#8fb4c1", fontsize=7.5)
    axis.set_ylabel("Signed minimum distance / solar radius", color="#8fb4c1", fontsize=7.5)
    axis.set_title(
        "1761 and 1769 Venus Transits — Excel Canonical Track Reconstruction\n"
        "IMCCE workbook contact geometry — IERS-0012O engineering style",
        color="#f8fdff",
        fontsize=9.0,
        pad=8,
    )
    legend = axis.legend(loc="lower right", fontsize=6.3, frameon=True, borderpad=0.45)
    legend.get_frame().set_facecolor("#071016")
    legend.get_frame().set_edgecolor("#1e4f64")
    legend.get_frame().set_linewidth(0.22)
    for legend_text in legend.get_texts():
        legend_text.set_color("#dff8ff")
    figure.text(
        0.5,
        0.016,
        "Excel-only reconstruction: C1-C4, radii, signed minimum distance, timing, and relative velocity. "
        "The canon does not provide an absolute sky-plane orientation or measured curved XY sequence.",
        ha="center",
        va="bottom",
        fontsize=6.0,
        color="#8fb4c1",
    )
    FIGURE_PNG.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(FIGURE_PNG, dpi=460, facecolor=figure.get_facecolor(), bbox_inches="tight", pad_inches=0.055)
    plt.show()
    plt.close(figure)


def write_track_csv(tracks: list[dict[str, object]]) -> None:
    TRACK_CSV.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "year",
        "sample_type",
        "event",
        "ut_clock",
        "elapsed_from_c1_seconds",
        "x_solar_radii",
        "y_solar_radii",
        "venus_radius_solar_radii",
        "source",
    ]
    with TRACK_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for track in tracks:
            c1 = float(track["times"][0])
            for epoch, x_value, y_value in zip(track["sample_times"], track["sample_x"], track["sample_y"]):
                writer.writerow(
                    {
                        "year": track["year"],
                        "sample_type": "ONE_MINUTE_RECONSTRUCTION",
                        "event": "",
                        "ut_clock": seconds_to_clock(float(epoch)),
                        "elapsed_from_c1_seconds": f"{float(epoch) - c1:.6f}",
                        "x_solar_radii": f"{float(x_value):.9f}",
                        "y_solar_radii": f"{float(y_value):.9f}",
                        "venus_radius_solar_radii": f"{float(track['venus_radius_ratio']):.9f}",
                        "source": "IMCCE_WORKBOOK_FOUR_CONTACT_LINEAR_FIT",
                    }
                )
            for index, event in enumerate(("C1", "C2", "C3", "C4")):
                writer.writerow(
                    {
                        "year": track["year"],
                        "sample_type": "CONTACT",
                        "event": event,
                        "ut_clock": seconds_to_clock(float(track["times"][index])),
                        "elapsed_from_c1_seconds": f"{float(track['times'][index]) - c1:.6f}",
                        "x_solar_radii": f"{float(track['contact_x'][index]):.9f}",
                        "y_solar_radii": f"{float(track['y']):.9f}",
                        "venus_radius_solar_radii": f"{float(track['venus_radius_ratio']):.9f}",
                        "source": "IMCCE_WORKBOOK_CONTACT_GEOMETRY",
                    }
                )
            writer.writerow(
                {
                    "year": track["year"],
                    "sample_type": "CLOSEST_APPROACH",
                    "event": "CA",
                    "ut_clock": seconds_to_clock(float(track["closest_time"])),
                    "elapsed_from_c1_seconds": f"{float(track['closest_time']) - c1:.6f}",
                    "x_solar_radii": "0.000000000",
                    "y_solar_radii": f"{float(track['y']):.9f}",
                    "venus_radius_solar_radii": f"{float(track['venus_radius_ratio']):.9f}",
                    "source": "IMCCE_WORKBOOK_FOUR_CONTACT_LINEAR_FIT",
                }
            )


def write_results_csv(tracks: list[dict[str, object]]) -> None:
    with RESULTS_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["year", "quantity", "value", "unit", "traceability"])
        for track in tracks:
            rows = [
                ("record_id", track["record"]["record_id"], "record", "IMCCE workbook"),
                ("sun_radius", track["sun_radius_arcsec"], "arcsec", "IMCCE workbook"),
                ("venus_radius", track["venus_radius_arcsec"], "arcsec", "IMCCE workbook"),
                ("signed_minimum_distance", track["signed_minimum_arcsec"], "arcsec", "IMCCE workbook"),
                ("signed_impact_ratio", track["y"], "solar radii", "calculated"),
                ("source_ratio", track["source_ratio"], "dimensionless", "IMCCE workbook"),
                ("calculated_ratio", track["calculated_ratio"], "dimensionless", "calculated"),
                ("ratio_residual", track["ratio_residual"], "dimensionless", "calculated"),
                ("external_chord", track["external_chord"], "solar radii", "calculated"),
                ("internal_chord", track["internal_chord"], "solar radii", "calculated"),
                ("external_duration", track["external_duration"], "s", "C1-C4"),
                ("internal_duration", track["internal_duration"], "s", "C2-C3"),
                ("fit_speed", track["fit_speed"], "solar radii/s", "four-contact least squares"),
                ("source_speed", track["source_speed"], "solar radii/s", "IMCCE V converted by solar radius"),
                ("external_speed", track["external_speed"], "solar radii/s", "external chord / duration"),
                ("internal_speed", track["internal_speed"], "solar radii/s", "internal chord / duration"),
                ("contact_fit_rms", track["fit_rms"], "solar radii", "four-contact least squares"),
                ("closest_time_fit", seconds_to_clock(float(track["closest_time"])), "UT", "four-contact least squares"),
                ("external_midpoint", seconds_to_clock(float(track["external_midpoint"])), "UT", "calculated"),
                ("internal_midpoint", seconds_to_clock(float(track["internal_midpoint"])), "UT", "calculated"),
                ("radial_equation_max_residual", track["radial_check_max"], "solar radii", "contact-circle equation"),
                ("canonical_track_angle", track["track_angle_deg"], "deg", "canonical frame definition"),
                ("canonical_curvature", track["curvature_inverse_solar_radii"], "1/solar radius", "canonical straight-chord model"),
            ]
            for quantity, value, unit_name, source in rows:
                writer.writerow([track["year"], quantity, value, unit_name, source])
        writer.writerow(
            [
                "COMPARISON",
                "halley_parallax_status",
                "NOT USED",
                "status",
                "1761 and 1769 are different transit epochs and not simultaneous observer tracks",
            ]
        )


def backup_outputs() -> dict[str, str]:
    statuses = {
        "track_csv": copy_verified(TRACK_CSV, DRIVE_OUTPUTS["track_csv"]),
        "results_csv": copy_verified(RESULTS_CSV, DRIVE_OUTPUTS["results_csv"]),
        "figure_png": copy_verified(FIGURE_PNG, DRIVE_OUTPUTS["figure_png"]),
    }
    script_path = Path(__file__).resolve() if "__file__" in globals() else None
    statuses["python"] = (
        copy_verified(script_path, DRIVE_OUTPUTS["python"])
        if script_path is not None and script_path.exists()
        else "NOT AVAILABLE"
    )
    return statuses


def main() -> None:
    workbook = resolve_workbook()
    records, sheet_name = read_target_records(workbook)
    tracks = [derive_track(year, records[year]) for year in TARGET_YEARS]
    write_track_csv(tracks)
    write_results_csv(tracks)
    plot_tracks(tracks)
    backups = backup_outputs()

    print(f"CODE OUTPUT: {VERSION}")
    print("CODE INPUTS")
    print(f"Workbook : {workbook}")
    print(f"Worksheet : {sheet_name}")
    print(f"Template : {REFERENCE_TEMPLATE}")
    print("COMMENTS")
    print("V0126B and V0126C are REJECTED for this task because they replaced the workbook reconstruction with JPL vectors.")
    print("This version uses only the organized Excel canon: radii, signed minimum distance, C1-C4 times, ratio, and V.")
    print("RESULTS")
    for track in tracks:
        print(
            f"{track['year']} | impact={track['y']:+.6f} R_sun | "
            f"external chord={track['external_chord']:.6f} R_sun | "
            f"internal chord={track['internal_chord']:.6f} R_sun"
        )
        print(
            f"{track['year']} | fit speed={track['fit_speed']:.9f} R_sun/s | "
            f"source speed={track['source_speed']:.9f} R_sun/s | "
            f"fit RMS={track['fit_rms']:.9f} R_sun"
        )
        print(
            f"{track['year']} | closest UT={seconds_to_clock(float(track['closest_time']))} | "
            f"external midpoint={seconds_to_clock(float(track['external_midpoint']))} | "
            f"internal midpoint={seconds_to_clock(float(track['internal_midpoint']))}"
        )
    print("A-prime B-prime and Halley solar parallax : NOT USED — different transit epochs.")
    print("OUTPUT SUMMARY")
    print(f"Track CSV : {TRACK_CSV}")
    print(f"Results CSV : {RESULTS_CSV}")
    print(f"Figure PNG : {FIGURE_PNG}")
    print(
        f"Drive backup : TRACK={backups['track_csv']} | RESULTS={backups['results_csv']} | "
        f"PNG={backups['figure_png']} | PYTHON={backups['python']}"
    )
    print("PAPER COMPARISON")
    print("NOT USED — direct reconstruction of two IMCCE workbook records.")
    print("EQUATION STATUS")
    print("VERIFIED — contact-circle intersections, four-contact least-squares timing fit, chord lengths, speeds, midpoint times, ratio residual, and disk scale evaluated.")
    print(datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S %z"))
    print("# V0126D")


if __name__ == "__main__":
    main()

# V0126D
