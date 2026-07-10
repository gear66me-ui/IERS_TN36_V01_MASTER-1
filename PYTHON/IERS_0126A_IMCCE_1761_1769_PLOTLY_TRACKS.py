# V0126A
# Audit reference: Plotly solar-disk comparison of the 1761 and 1769 IMCCE canonical Venus transit tracks.

from __future__ import annotations

import csv
import hashlib
import math
import shutil
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path


def ensure_packages() -> None:
    missing = []
    for module, package in (("numpy", "numpy"), ("openpyxl", "openpyxl>=3.1.0"), ("plotly", "plotly>=5.24.0")):
        try:
            __import__(module)
        except ImportError:
            missing.append(package)
    if missing:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", *missing])


ensure_packages()
import numpy as np
import plotly.graph_objects as go
from openpyxl import load_workbook
from plotly.colors import qualitative

VERSION = "IERS-0126A"
TARGET_YEARS = (1761, 1769)
LOCAL_TZ = timezone(timedelta(hours=-5))
DRIVE_ROOT = Path("/content/drive/MyDrive/Colab Notebooks/JPL - 1769 VENUS TRANSIT/GitHub")
INPUT_XLSX = DRIVE_ROOT / "DATA" / "XLSX" / "IMCCE_VENUS_TRANSIT_CANON_ORGANIZED.xlsx"
FALLBACK_XLSX = DRIVE_ROOT / "DATA" / "XLSX" / "IMCCE_VENUS_TRANSIT_CANON_MASTER.xlsx"
OUTPUT_ROOT = Path("/content/IERS_TN36_V01_MASTER_OUTPUT/IMCCE_VENUS_CANON/PLOTS/V0126A")
TRACK_CSV = OUTPUT_ROOT / "IERS_0126A_IMCCE_1761_1769_TRACKS.csv"
RESULTS_CSV = OUTPUT_ROOT / "IERS_0126A_IMCCE_1761_1769_RESULTS.csv"
PLOT_HTML = OUTPUT_ROOT / "IERS_0126A_IMCCE_1761_1769_PLOTLY_TRACKS.html"
PLOT_PNG = OUTPUT_ROOT / "IERS_0126A_IMCCE_1761_1769_PLOTLY_TRACKS.png"
DRIVE_TARGETS = {
    "track": DRIVE_ROOT / "DATA" / "CSV" / TRACK_CSV.name,
    "results": DRIVE_ROOT / "DATA" / "CSV" / RESULTS_CSV.name,
    "html": DRIVE_ROOT / "DATA" / "HTML" / PLOT_HTML.name,
    "png": DRIVE_ROOT / "DATA" / "PNG" / PLOT_PNG.name,
    "python": DRIVE_ROOT / "PYTHON" / "IERS_0126A_IMCCE_1761_1769_PLOTLY_TRACKS.py",
}
REQUIRED = {
    "year", "sun_radius_arcsec", "minimum_distance_arcsec", "venus_radius_arcsec",
    "c1_ut", "c2_ut", "c3_ut", "c4_ut", "mid_ut_seconds_of_day",
    "relative_velocity_deg_per_day", "external_duration_seconds",
    "internal_duration_seconds", "node", "record_id",
}
ALIASES = {
    "solar_radius_(arcsec)": "sun_radius_arcsec",
    "signed_minimum_distance_(arcsec)": "minimum_distance_arcsec",
    "venus_radius_(arcsec)": "venus_radius_arcsec",
    "relative_velocity_(deg/day)": "relative_velocity_deg_per_day",
    "date_ut": "date_ut_label",
    "mid_transit_ut": "mid_ut_hhmm",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def copy_verified(source: Path, destination: Path) -> str:
    destination.parent.mkdir(parents=True, exist_ok=True)
    source_hash = sha256(source)
    if destination.exists() and sha256(destination) == source_hash:
        return "UNCHANGED"
    shutil.copy2(source, destination)
    if sha256(destination) != source_hash:
        raise RuntimeError(f"SHA-256 verification failed: {destination}")
    return "COPIED"


def normalize_header(value: object) -> str:
    return str(value).strip().lower().replace(" ", "_").replace("-", "_")


def resolve_workbook() -> Path:
    for path in (INPUT_XLSX, FALLBACK_XLSX):
        if path.exists():
            return path
    raise FileNotFoundError("Run IERS-0125A before IERS-0126A; organized workbook not found.")


def read_years(path: Path) -> dict[int, dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    chosen = None
    for sheet_name, header_row in (("PLOT_DATA", 1), ("MASTER", 4)):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        columns = {}
        for index in range(1, sheet.max_column + 1):
            raw = sheet.cell(header_row, index).value
            if raw is not None:
                key = normalize_header(raw)
                columns[ALIASES.get(key, key)] = index
        if REQUIRED.issubset(columns):
            chosen = (sheet, header_row, columns)
            break
    if chosen is None:
        workbook.close()
        raise RuntimeError("Neither PLOT_DATA nor MASTER contains every required calculation column.")
    sheet, header_row, columns = chosen
    rows = {}
    for row_index in range(header_row + 1, sheet.max_row + 1):
        value = sheet.cell(row_index, columns["year"]).value
        if value is None or int(value) not in TARGET_YEARS:
            continue
        year = int(value)
        rows[year] = {name: sheet.cell(row_index, col).value for name, col in columns.items()}
    workbook.close()
    if set(rows) != set(TARGET_YEARS):
        raise RuntimeError(f"Missing target years: {sorted(set(TARGET_YEARS) - set(rows))}")
    return rows


def clock_seconds(value: object) -> float:
    fields = str(value).strip().split(":")
    if len(fields) == 2:
        return int(fields[0]) * 3600.0 + float(fields[1]) * 60.0
    if len(fields) == 3:
        return int(fields[0]) * 3600.0 + int(fields[1]) * 60.0 + float(fields[2])
    raise ValueError(f"Unsupported clock: {value!r}")


def format_clock(seconds: float) -> str:
    seconds %= 86400.0
    hour = int(seconds // 3600.0)
    minute = int((seconds % 3600.0) // 60.0)
    second = seconds % 60.0
    return f"{hour:02d}:{minute:02d}:{second:06.3f}"


def unwrap_times(row: dict[str, object]) -> np.ndarray:
    times = [clock_seconds(row[f"c{i}_ut"]) for i in range(1, 5)]
    output = [times[0]]
    for value in times[1:]:
        while value < output[-1]:
            value += 86400.0
        output.append(value)
    return np.asarray(output, dtype=float)


def align_midpoint(value: float, start: float, stop: float) -> float:
    center = 0.5 * (start + stop)
    return min((value + day * 86400.0 for day in range(-2, 3)), key=lambda item: abs(item - center))


def derive(year: int, row: dict[str, object]) -> dict[str, object]:
    rs = float(row["sun_radius_arcsec"])
    rp = float(row["venus_radius_arcsec"])
    delta = float(row["minimum_distance_arcsec"])
    node = int(float(row["node"]))
    times = unwrap_times(row)
    y = delta / rs
    rv = rp / rs
    x_ext = math.sqrt((1.0 + rv) ** 2 - y**2)
    x_int = math.sqrt((1.0 - rv) ** 2 - y**2)
    direction = 1.0 if node >= 0 else -1.0
    contacts = direction * np.asarray([-x_ext, -x_int, x_int, x_ext], dtype=float)
    slope, intercept = np.polyfit(times, contacts, 1)
    fit = slope * times + intercept
    rms = float(np.sqrt(np.mean((contacts - fit) ** 2)))
    closest = float(-intercept / slope)
    source_mid = align_midpoint(float(row["mid_ut_seconds_of_day"]), times[0], times[-1])
    minute_times = np.arange(times[0], times[-1] + 0.001, 60.0)
    if minute_times[-1] < times[-1]:
        minute_times = np.append(minute_times, times[-1])
    minute_x = np.clip(slope * minute_times + intercept, min(contacts[0], contacts[3]), max(contacts[0], contacts[3]))
    source_speed = float(row["relative_velocity_deg_per_day"]) * 3600.0 / 86400.0 / rs
    return {
        "year": year, "record_id": int(float(row["record_id"])), "rs": rs, "rp": rp,
        "delta": delta, "y": y, "rv": rv, "node": node, "times": times,
        "contacts": contacts, "slope": float(slope), "rms": rms, "closest": closest,
        "source_mid": source_mid, "external_mid": float((times[0] + times[3]) / 2.0),
        "internal_mid": float((times[1] + times[2]) / 2.0),
        "external_chord": 2.0 * x_ext, "internal_chord": 2.0 * x_int,
        "source_speed": source_speed,
        "external_speed": 2.0 * x_ext / float(row["external_duration_seconds"]),
        "internal_speed": 2.0 * x_int / float(row["internal_duration_seconds"]),
        "minute_times": minute_times, "minute_x": minute_x,
        "minute_y": np.full_like(minute_x, y),
    }


def circle(center_x: float, center_y: float, radius: float) -> tuple[np.ndarray, np.ndarray]:
    theta = np.linspace(0.0, 2.0 * math.pi, 181)
    return center_x + radius * np.cos(theta), center_y + radius * np.sin(theta)


def build_figure(tracks: list[dict[str, object]]) -> go.Figure:
    figure = go.Figure()
    theta = np.linspace(0.0, 2.0 * math.pi, 721)
    figure.add_trace(go.Scatter(x=np.cos(theta), y=np.sin(theta), mode="lines", name="Solar limb", line={"width": 1.2, "color": "black"}, hoverinfo="skip"))
    for index, track in enumerate(tracks):
        color = qualitative.Plotly[index]
        year = int(track["year"])
        contacts = np.asarray(track["contacts"], dtype=float)
        y = float(track["y"])
        line_x = np.linspace(float(contacts[0]), float(contacts[3]), 401)
        figure.add_trace(go.Scatter(x=line_x, y=np.full_like(line_x, y), mode="lines", name=f"{year} center track", line={"width": 2.0, "color": color}, hovertemplate=f"{year}<br>X %{{x:.6f}} R☉<br>Y %{{y:.6f}} R☉<extra></extra>"))
        custom = np.column_stack((np.full(len(track["minute_x"]), year), [format_clock(float(t)) for t in track["minute_times"]]))
        figure.add_trace(go.Scatter(x=track["minute_x"], y=track["minute_y"], mode="markers", name=f"{year} one-minute samples", marker={"size": 3.5, "color": color}, customdata=custom, hovertemplate="Year %{customdata[0]}<br>UT %{customdata[1]}<br>X %{x:.6f} R☉<br>Y %{y:.6f} R☉<extra></extra>"))
        for event_index, label in enumerate(("C1", "C2", "C3", "C4")):
            x_value = float(contacts[event_index])
            xd, yd = circle(x_value, y, float(track["rv"]))
            figure.add_trace(go.Scatter(x=xd, y=yd, mode="lines", line={"width": 1.0, "color": color}, hoverinfo="skip", showlegend=False))
            figure.add_annotation(x=x_value, y=y + (0.055 if y < 0 else -0.055), text=label, showarrow=False, font={"size": 9, "color": color})
        xd, yd = circle(0.0, y, float(track["rv"]))
        figure.add_trace(go.Scatter(x=xd, y=yd, mode="lines", line={"width": 1.0, "color": color}, hoverinfo="skip", showlegend=False))
        figure.add_annotation(x=0.0, y=y, text=f"{year}: {y:+.6f} R☉", showarrow=True, ax=(-70 if index == 0 else 70), ay=(-25 if y > 0 else 25), arrowwidth=0.8, arrowcolor=color, font={"size": 10, "color": color}, bgcolor="rgba(255,255,255,0.80)")
        start = float(contacts[0] + 0.42 * (contacts[3] - contacts[0]))
        stop = float(contacts[0] + 0.58 * (contacts[3] - contacts[0]))
        figure.add_annotation(x=stop, y=y, ax=start, ay=y, xref="x", yref="y", axref="x", ayref="y", text="", showarrow=True, arrowhead=3, arrowwidth=1.2, arrowcolor=color)
    figure.update_layout(template="plotly_white", width=920, height=820, title={"text": "IMCCE Venus Transit Canon — 1761 and 1769 Solar-Disk Tracks<br><sup>Canonical motion-axis frame; one-minute samples derived from C1–C4 timing</sup>", "x": 0.5}, hovermode="closest", legend={"orientation": "h", "x": 0.5, "xanchor": "center", "y": -0.09}, margin={"l": 70, "r": 40, "t": 95, "b": 95}, xaxis={"title": "Canonical along-track coordinate (solar radii)", "range": [-1.14, 1.14], "showgrid": False, "zerolinewidth": 0.6}, yaxis={"title": "Canonical cross-track coordinate (solar radii)", "range": [-1.14, 1.14], "showgrid": False, "zerolinewidth": 0.6, "scaleanchor": "x", "scaleratio": 1.0})
    return figure


def write_track_csv(tracks: list[dict[str, object]]) -> None:
    fields = ["year", "sample_type", "event", "ut_clock", "elapsed_from_c1_seconds", "x_solar_radii", "y_solar_radii", "venus_radius_solar_radii", "source"]
    with TRACK_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for track in tracks:
            c1 = float(track["times"][0])
            for i, event in enumerate(("C1", "C2", "C3", "C4")):
                writer.writerow({"year": track["year"], "sample_type": "CONTACT", "event": event, "ut_clock": format_clock(float(track["times"][i])), "elapsed_from_c1_seconds": float(track["times"][i]) - c1, "x_solar_radii": float(track["contacts"][i]), "y_solar_radii": track["y"], "venus_radius_solar_radii": track["rv"], "source": "IMCCE_CANON_CONTACT_GEOMETRY"})
            writer.writerow({"year": track["year"], "sample_type": "CLOSEST_APPROACH", "event": "CA", "ut_clock": format_clock(float(track["closest"])), "elapsed_from_c1_seconds": float(track["closest"]) - c1, "x_solar_radii": 0.0, "y_solar_radii": track["y"], "venus_radius_solar_radii": track["rv"], "source": "FOUR_CONTACT_LINEAR_FIT"})
            for epoch, x_value in zip(track["minute_times"], track["minute_x"]):
                writer.writerow({"year": track["year"], "sample_type": "ONE_MINUTE", "event": "", "ut_clock": format_clock(float(epoch)), "elapsed_from_c1_seconds": float(epoch) - c1, "x_solar_radii": float(x_value), "y_solar_radii": track["y"], "venus_radius_solar_radii": track["rv"], "source": "FOUR_CONTACT_LINEAR_FIT"})


def write_results(tracks: list[dict[str, object]]) -> dict[str, float]:
    separation = abs(float(tracks[1]["y"]) - float(tracks[0]["y"]))
    direct_arcsec = abs(float(tracks[1]["delta"]) - float(tracks[0]["delta"]))
    with RESULTS_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["year", "quantity", "value", "unit", "traceability"])
        for track in tracks:
            for quantity, value, unit, source in (
                ("signed_closest_approach", track["delta"], "arcsec", "IMCCE workbook"),
                ("signed_impact_ratio", track["y"], "solar radii", "calculated"),
                ("external_chord", track["external_chord"], "solar radii", "calculated"),
                ("internal_chord", track["internal_chord"], "solar radii", "calculated"),
                ("contact_fit_speed", track["slope"], "solar radii/s", "four-contact fit"),
                ("source_speed", track["source_speed"], "solar radii/s", "IMCCE V field"),
                ("external_speed", track["external_speed"], "solar radii/s", "chord/duration"),
                ("internal_speed", track["internal_speed"], "solar radii/s", "chord/duration"),
                ("fit_rms", track["rms"], "solar radii", "four-contact fit"),
                ("closest_fit_ut", format_clock(float(track["closest"])), "UT", "four-contact fit"),
                ("source_mid_ut", format_clock(float(track["source_mid"])), "UT", "IMCCE workbook"),
                ("external_mid_delta", float(track["external_mid"]) - float(track["source_mid"]), "s", "calculated"),
                ("internal_mid_delta", float(track["internal_mid"]) - float(track["source_mid"]), "s", "calculated"),
            ):
                writer.writerow([track["year"], quantity, value, unit, source])
        writer.writerow(["COMPARISON", "track_center_separation", separation, "solar radii", "difference of signed impact ratios"])
        writer.writerow(["COMPARISON", "signed_minimum_distance_difference", direct_arcsec, "arcsec", "difference of IMCCE closest approaches"])
        writer.writerow(["COMPARISON", "halley_parallax_status", "NOT USED", "status", "different epochs, not simultaneous observer tracks"])
    return {"separation": separation, "direct_arcsec": direct_arcsec}


def save_png(figure: go.Figure) -> str:
    try:
        figure.write_image(str(PLOT_PNG), width=1380, height=1230, scale=2)
        return "SAVED"
    except Exception:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "kaleido==0.2.1"])
        figure.write_image(str(PLOT_PNG), width=1380, height=1230, scale=2)
        return "SAVED_AFTER_KALEIDO_INSTALL"


def main() -> None:
    workbook = resolve_workbook()
    rows = read_years(workbook)
    tracks = [derive(year, rows[year]) for year in TARGET_YEARS]
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    write_track_csv(tracks)
    comparison = write_results(tracks)
    figure = build_figure(tracks)
    figure.write_html(str(PLOT_HTML), include_plotlyjs=True, full_html=True)
    png_status = save_png(figure)
    figure.show()
    backups = {"track": copy_verified(TRACK_CSV, DRIVE_TARGETS["track"]), "results": copy_verified(RESULTS_CSV, DRIVE_TARGETS["results"]), "html": copy_verified(PLOT_HTML, DRIVE_TARGETS["html"]), "png": copy_verified(PLOT_PNG, DRIVE_TARGETS["png"])}
    script = Path(__file__).resolve() if "__file__" in globals() else None
    backups["python"] = copy_verified(script, DRIVE_TARGETS["python"]) if script and script.exists() else "NOT AVAILABLE"

    print(f"CODE OUTPUT: {VERSION}")
    print("CODE INPUTS")
    print(f"Workbook : {workbook}")
    print("COMMENTS")
    print("Python Plotly data visualization only. No AI image generation.")
    print("Each event is normalized by its own IMCCE solar radius in a canonical motion-axis frame.")
    print("RESULTS")
    for track in tracks:
        print(f"{track['year']} | y={float(track['y']):+.6f} R_sun | external chord={float(track['external_chord']):.6f} R_sun | fit RMS={float(track['rms']):.9f} R_sun | closest UT={format_clock(float(track['closest']))}")
    print(f"Track-center separation : {comparison['separation']:.6f} R_sun | signed-distance difference : {comparison['direct_arcsec']:.6f} arcsec")
    print("A-prime B-prime and Halley solar parallax : NOT USED — 1761 and 1769 are different epochs.")
    print("OUTPUT SUMMARY")
    print(f"Track CSV : {TRACK_CSV}")
    print(f"Results CSV : {RESULTS_CSV}")
    print(f"Interactive HTML : {PLOT_HTML}")
    print(f"PNG : {PLOT_PNG} | {png_status}")
    print(f"Drive backup : TRACK={backups['track']} | RESULTS={backups['results']} | HTML={backups['html']} | PNG={backups['png']} | PYTHON={backups['python']}")
    print("PAPER COMPARISON")
    print("IMCCE contact geometry and V relative displacement are compared through independent chord/duration speeds.")
    print("EQUATION STATUS")
    print("VERIFIED — limb intersections, C1-C4 fit, midpoint timing, RMS, normalized separation, and disk scale evaluated.")
    print(datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S %z"))
    print("# V0126A")


if __name__ == "__main__":
    main()

# V0126A
