# V0122E
# Audit reference: final integrity audit and Colab delivery for IMCCE Venus Transit Canon CSV and XLSX outputs.

from __future__ import annotations

import csv
import hashlib
import json
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl>=3.1.0"])
    from openpyxl import load_workbook

VERSION = "IERS-0122E"
OUTPUT_ROOT = Path("/content/IERS_TN36_V01_MASTER_OUTPUT/IMCCE_VENUS_CANON")
MASTER_CSV = OUTPUT_ROOT / "IMCCE_VENUS_TRANSIT_CANON_MASTER.csv"
MASTER_XLSX = OUTPUT_ROOT / "IMCCE_VENUS_TRANSIT_CANON_MASTER.xlsx"
AUDIT_JSON = OUTPUT_ROOT / "IERS_0122E_FINAL_DELIVERY_AUDIT.json"
LOCAL_TZ = timezone(timedelta(hours=-5))
TARGET_YEARS = (1761, 1769, 1874, 1882, 2004, 2012)
EXPECTED_SHEETS = ["README", "MASTER", "1761", "1769", "1874", "1882", "2004", "2012"]
EXPECTED_ROWS = 77
AUTO_DOWNLOAD = True


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_csv() -> tuple[list[str], list[dict[str, str]]]:
    if not MASTER_CSV.exists():
        raise FileNotFoundError(f"Missing final CSV: {MASTER_CSV}")
    with MASTER_CSV.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or [])
        rows = list(reader)
    if len(rows) != EXPECTED_ROWS:
        raise RuntimeError(f"CSV row count mismatch: expected {EXPECTED_ROWS}, found {len(rows)}")
    if len(headers) != len(set(headers)):
        raise RuntimeError("CSV contains duplicate column names")
    return headers, rows


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return format(value, ".15g")
    return str(value)


def verify_master_sheet(headers: list[str], csv_rows: list[dict[str, str]], ws) -> None:
    workbook_headers = [normalize_cell(ws.cell(row=4, column=index).value) for index in range(1, len(headers) + 1)]
    if workbook_headers != headers:
        raise RuntimeError("MASTER workbook headers do not match CSV headers")
    if ws.max_row != EXPECTED_ROWS + 4:
        raise RuntimeError(f"MASTER worksheet row count mismatch: {ws.max_row}")

    key_columns = ["record_id", "year", "jd_tdb", "record_status", "ratio_validation"]
    key_indices = {name: headers.index(name) + 1 for name in key_columns}
    for offset, csv_row in enumerate(csv_rows, start=5):
        for name, column in key_indices.items():
            csv_value = normalize_cell(csv_row[name])
            xlsx_value = normalize_cell(ws.cell(row=offset, column=column).value)
            if name in {"record_id", "year"}:
                csv_value = normalize_cell(int(float(csv_value)))
            elif name == "jd_tdb":
                csv_value = normalize_cell(float(csv_value))
            if csv_value != xlsx_value:
                raise RuntimeError(
                    f"CSV/XLSX mismatch at MASTER row {offset}, column {name}: "
                    f"CSV={csv_value!r}, XLSX={xlsx_value!r}"
                )


def verify_target_sheets(headers: list[str], csv_rows: list[dict[str, str]], workbook) -> None:
    year_column = headers.index("year") + 1
    record_column = headers.index("record_id") + 1
    csv_by_year = {int(row["year"]): int(row["record_id"]) for row in csv_rows}
    for year in TARGET_YEARS:
        ws = workbook[str(year)]
        if ws.max_row != 5:
            raise RuntimeError(f"Worksheet {year} must contain exactly one data row")
        workbook_year = int(ws.cell(row=5, column=year_column).value)
        workbook_record = int(ws.cell(row=5, column=record_column).value)
        if workbook_year != year or workbook_record != csv_by_year[year]:
            raise RuntimeError(
                f"Worksheet {year} mismatch: year={workbook_year}, record={workbook_record}, "
                f"expected record={csv_by_year[year]}"
            )


def verify_xlsx(headers: list[str], csv_rows: list[dict[str, str]]) -> dict[str, object]:
    if not MASTER_XLSX.exists():
        raise FileNotFoundError(f"Missing final workbook: {MASTER_XLSX}")
    workbook = load_workbook(MASTER_XLSX, data_only=False, read_only=False)
    if workbook.sheetnames != EXPECTED_SHEETS:
        raise RuntimeError(f"Workbook sheet order mismatch: {workbook.sheetnames}")
    verify_master_sheet(headers, csv_rows, workbook["MASTER"])
    verify_target_sheets(headers, csv_rows, workbook)

    formula_count = 0
    error_count = 0
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                if value in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}:
                    error_count += 1
    workbook.close()
    if error_count:
        raise RuntimeError(f"Workbook contains {error_count} spreadsheet error values")
    return {"sheet_count": len(EXPECTED_SHEETS), "formula_count": formula_count, "error_count": error_count}


def build_audit(headers: list[str], rows: list[dict[str, str]], xlsx_audit: dict[str, object]) -> dict[str, object]:
    years = [int(row["year"]) for row in rows]
    complete = sum(row["record_status"] == "COMPLETE" for row in rows)
    header_only = sum(row["record_status"] == "SOURCE_HEADER_ONLY" for row in rows)
    ratio_review = sum(row["ratio_validation"] == "REVIEW" for row in rows)
    return {
        "version": VERSION,
        "audited_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "status": "PASS",
        "csv": {
            "path": str(MASTER_CSV),
            "bytes": MASTER_CSV.stat().st_size,
            "sha256": sha256(MASTER_CSV),
            "rows": len(rows),
            "columns": len(headers),
        },
        "xlsx": {
            "path": str(MASTER_XLSX),
            "bytes": MASTER_XLSX.stat().st_size,
            "sha256": sha256(MASTER_XLSX),
            **xlsx_audit,
        },
        "records": {
            "complete": complete,
            "source_header_only": header_only,
            "ratio_review": ratio_review,
            "minimum_year": min(years),
            "maximum_year": max(years),
            "target_years_verified": list(TARGET_YEARS),
        },
    }


def download_outputs() -> str:
    if not AUTO_DOWNLOAD:
        return "DISABLED"
    try:
        from google.colab import files
    except ImportError:
        return "NOT COLAB"
    files.download(str(MASTER_CSV))
    files.download(str(MASTER_XLSX))
    return "REQUESTED"


def main() -> None:
    headers, rows = load_csv()
    xlsx_audit = verify_xlsx(headers, rows)
    audit = build_audit(headers, rows, xlsx_audit)
    AUDIT_JSON.write_text(json.dumps(audit, indent=2) + "\n", encoding="utf-8")
    download_status = download_outputs()

    print(f"CODE OUTPUT: {VERSION}")
    print("CODE INPUTS")
    print(f"CSV : {MASTER_CSV}")
    print(f"XLSX : {MASTER_XLSX}")
    print("COMMENTS")
    print("Final CSV/XLSX consistency, workbook structure, target-year sheets, hashes, and spreadsheet errors audited.")
    print("RESULTS")
    print(f"Status : PASS | Rows : {audit['csv']['rows']} | Columns : {audit['csv']['columns']} | Sheets : {audit['xlsx']['sheet_count']}")
    print(f"CSV SHA256 : {audit['csv']['sha256']}")
    print(f"XLSX SHA256 : {audit['xlsx']['sha256']}")
    print("OUTPUT SUMMARY")
    print(f"CSV : {MASTER_CSV}")
    print(f"XLSX : {MASTER_XLSX}")
    print(f"Audit : {AUDIT_JSON}")
    print(f"Colab downloads : {download_status}")
    print("PAPER COMPARISON")
    print("NOT USED — final integrity and delivery stage.")
    print("EQUATION STATUS")
    print("VERIFIED — CSV/XLSX key fields and six target-year records agree exactly.")
    print(datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S %z"))
    print("# V0122E")


if __name__ == "__main__":
    main()

# V0122E
